from fastapi import FastAPI, APIRouter, HTTPException, Depends, Response, Request, status
from fastapi.responses import JSONResponse, StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import aiohttp
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

# Import AI helpers
from ai_helpers import (
    analyze_vendor_scoring,
    analyze_tender_proposal,
    analyze_contract_classification,
    analyze_po_items,
    match_invoice_to_milestone
)

# Import all models from the models package
from models import (
    User, UserSession, UserRole,
    Vendor, VendorType, VendorStatus, RiskCategory,
    Tender, TenderStatus, EvaluationCriteria, Proposal, ProposalStatus,
    Contract, ContractStatus,
    Invoice, InvoiceStatus,
    PurchaseOrder, POItem, POStatus,
    Resource, ResourceStatus, WorkType, Relative,
    Asset, AssetStatus, AssetCondition, Building, Floor, AssetCategory,
    OSR, OSRType, OSRCategory, OSRStatus, OSRPriority,
    Notification, AuditLog
)

# Import utilities
from utils.database import db, client
from utils.auth import hash_password, verify_password, get_current_user, require_auth, require_role
from utils.helpers import generate_number, determine_outsourcing_classification, determine_noc_requirement

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")


# ==================== HELPER FUNCTIONS ====================
def calculate_vendor_registration_score(vendor_data: dict) -> dict:
    """
    Calculate vendor registration score based on 15 Yes/No questions.
    1 point for each "Yes" answer.
    Returns: dict with score, percentage, and risk_category
    """
    total_questions = 15
    score = 0
    
    # All questions award 1 point for "Yes"
    if vendor_data.get('vat_number'): score += 1
    if vendor_data.get('unified_number'): score += 1
    if vendor_data.get('cr_number'): score += 1
    if vendor_data.get('cr_expiry_date'): score += 1
    if vendor_data.get('cr_country_city'): score += 1
    if vendor_data.get('license_number'): score += 1
    if vendor_data.get('license_expiry_date'): score += 1
    if vendor_data.get('activity_description'): score += 1
    if vendor_data.get('number_of_employees'): score += 1
    if vendor_data.get('country_list'): score += 1
    if vendor_data.get('financial_details'): score += 1
    if vendor_data.get('branches_subsidiaries'): score += 1
    if vendor_data.get('key_customers'): score += 1
    if vendor_data.get('financial_statements_2years'): score += 1
    if vendor_data.get('documents_pdf_attached'): score += 1
    
    percentage = (score / total_questions) * 100
    
    # Determine risk category based on percentage
    if percentage >= 80:
        risk_category = 'low'
    elif percentage >= 60:
        risk_category = 'medium'
    elif percentage >= 40:
        risk_category = 'high'
    else:
        risk_category = 'very_high'
    
    return {
        'score': score,
        'total': total_questions,
        'percentage': round(percentage, 2),
        'risk_category': risk_category
    }

def calculate_due_diligence_score(vendor_data: dict) -> dict:
    """
    Calculate due diligence score based on 60+ Yes/No questions.
    1 point for each "Yes" answer (some questions are reverse-scored).
    Returns: dict with score, percentage, and risk_category
    """
    score = 0
    total_questions = 64
    
    # POSITIVE SCORING (Yes = +1 point) - Good practices
    positive_fields = [
        'dd_location_moved_closed',  # Has alternative locations
        'dd_vendor_opened_branches',  # Growth indicator
        'dd_bc_alternative_locations',
        'dd_bc_test_continuity_regularly',
        'dd_bc_certified_standard',
        'dd_bc_dedicated_staff',
        'dd_bc_risk_assessment_done',
        'dd_bc_essential_activities_identified',
        'dd_bc_strategy_exists',
        'dd_bc_emergency_responders',
        'dd_bc_update_arrangements',
        'dd_bc_exercising_strategy',
        'dd_bc_evidence_of_tests',
        'dd_bc_test_results_improve',
        'dd_bc_management_trained',
        'dd_bc_staff_aware',
        'dd_bc_it_continuity_plan',
        'dd_bc_data_backed_up_offsite',
        'dd_bc_vital_documents_backed_up',
        'dd_bc_critical_suppliers_identified',
        'dd_bc_coordinated_with_suppliers',
        'dd_bc_communication_method',
        'dd_bc_pr_crisis_management',
        'dd_fraud_whistleblowing_mechanism',
        'dd_fraud_prevention_procedures',
        'dd_op_documented_procedures',
        'dd_op_internal_audit',
        'dd_op_coi_policies',
        'dd_op_complaint_handling',
        'dd_op_insurance_contracts',
        'dd_cyber_security_procedures',
        'dd_safety_security_24_7',
        'dd_safety_cctv_equipment',
        'dd_safety_fire_exits_equipment',
        'dd_hr_localization_policy',
        'dd_hr_hiring_policy',
        'dd_hr_background_investigation',
        'dd_hr_academic_verification',
        'dd_op_financial_statements_audited',
        'dd_data_management_policy',
        'dd_sama_consumer_protection_understanding',
        'dd_sama_consumer_protection_compliance',
    ]
    
    # NEGATIVE SCORING (Yes = -1 point) - Red flags
    negative_fields = [
        'dd_ownership_change_last_year',  # Instability
        'dd_bc_exposed_to_events',  # Past disruptions
        'dd_conflicts_of_interest',
        'dd_bc_rely_on_third_parties',  # Dependency risk
        'dd_bc_plan_to_subcontract',
        'dd_fraud_internal_last_year',
        'dd_fraud_burglary_theft_last_year',
        'dd_op_criminal_cases_last_3years',
        'dd_op_customer_complaints_last_year',
        'dd_cloud_services',  # Potential risk
        'dd_cyber_data_outside_ksa',
        'dd_cyber_remote_access_outside_ksa',
        'dd_cyber_digital_channel_services',
        'dd_cyber_card_payment_services',
        'dd_cyber_third_party_access',
        'dd_op_legal_public_representation',
        'dd_op_activities_regulated',
        'dd_related_party_to_bank',
    ]
    
    # NEUTRAL (Yes/No both acceptable) - Informational only
    neutral_fields = [
        'dd_op_specific_license_required',
        'dd_op_services_outside_ksa',
    ]
    
    # Calculate positive points
    for field in positive_fields:
        if vendor_data.get(field) is True:
            score += 1
    
    # Calculate negative points (reverse scoring)
    for field in negative_fields:
        if vendor_data.get(field) is False:  # No red flag = good
            score += 1
    
    percentage = (score / total_questions) * 100
    
    # Determine risk category based on percentage
    if percentage >= 75:
        risk_category = 'low'
    elif percentage >= 50:
        risk_category = 'medium'
    elif percentage >= 30:
        risk_category = 'high'
    else:
        risk_category = 'very_high'
    
    return {
        'score': score,
        'total': total_questions,
        'percentage': round(percentage, 2),
        'risk_category': risk_category
    }

def calculate_dd_risk_adjustment(vendor_data: dict) -> float:
    """
    Calculate risk score adjustment based on Due Diligence responses.
    Uses new Yes/No scoring system.
    Returns risk score from 0-100.
    """
    dd_result = calculate_due_diligence_score(vendor_data)
    
    # Convert percentage to risk score (inverse)
    # Higher percentage = lower risk score
    risk_score = 100 - dd_result['percentage']
    
    return risk_score


# ==================== REQUEST/RESPONSE MODELS ====================
class LoginRequest(BaseModel):
    email: EmailStr
    password: str

class RegisterRequest(BaseModel):
    email: EmailStr
    password: str
    name: str
    role: Optional[UserRole] = UserRole.PROCUREMENT_OFFICER

class ProposalEvaluationRequest(BaseModel):
    """Request model for evaluating proposals"""
    vendor_reliability_stability: float
    delivery_warranty_backup: float
    technical_experience: float
    cost_score: float
    meets_requirements: float

# ==================== AUTH ENDPOINTS ====================
@api_router.post("/auth/register")
async def register(register_data: RegisterRequest):
    """Register a new user (admin only endpoint for creating users)"""
    # Check if user already exists
    existing_user = await db.users.find_one({"email": register_data.email})
    if existing_user:
        raise HTTPException(status_code=400, detail="User already exists")
    
    # Create new user
    user = User(
        email=register_data.email,
        name=register_data.name,
        password=hash_password(register_data.password),
        role=register_data.role
    )
    
    user_doc = user.model_dump()
    user_doc["created_at"] = user_doc["created_at"].isoformat()
    await db.users.insert_one(user_doc)
    
    # Remove password from response
    user_dict = user.model_dump()
    user_dict.pop('password', None)
    
    return {"message": "User created successfully", "user": user_dict}

@api_router.post("/auth/login")
async def login(login_data: LoginRequest, response: Response):
    """Login with email and password"""
    # Find user
    user_doc = await db.users.find_one({"email": login_data.email})
    if not user_doc:
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    # Verify password
    if not verify_password(login_data.password, user_doc.get("password", "")):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    # Convert datetime strings
    if isinstance(user_doc.get('created_at'), str):
        user_doc['created_at'] = datetime.fromisoformat(user_doc['created_at'])
    
    user = User(**user_doc)
    
    # Create session
    session_token = str(uuid.uuid4()) + str(uuid.uuid4())
    expires_at = datetime.now(timezone.utc) + timedelta(days=7)
    
    session = UserSession(
        user_id=user.id,
        session_token=session_token,
        expires_at=expires_at
    )
    
    session_doc = session.model_dump()
    session_doc["expires_at"] = session_doc["expires_at"].isoformat()
    session_doc["created_at"] = session_doc["created_at"].isoformat()
    await db.user_sessions.insert_one(session_doc)
    
    # Set cookie
    response.set_cookie(
        key="session_token",
        value=session_token,
        httponly=True,
        secure=True,
        samesite="none",  # Changed to 'none' for cross-origin requests
        path="/",
        max_age=7 * 24 * 60 * 60
    )
    
    # Remove password from response
    user_dict = user.model_dump()
    user_dict.pop('password', None)
    
    return {"user": user_dict}

@api_router.post("/auth/auto-login")
async def auto_login(response: Response):
    """
    Auto-login endpoint that creates a session for the default user without credentials.
    This allows users to access the application without manual login.
    """
    # Use default procurement officer account
    default_email = "procurement@test.com"
    
    # Find default user
    user_doc = await db.users.find_one({"email": default_email})
    if not user_doc:
        raise HTTPException(status_code=500, detail="Default user not found")
    
    # Convert datetime strings
    if isinstance(user_doc.get('created_at'), str):
        user_doc['created_at'] = datetime.fromisoformat(user_doc['created_at'])
    
    user = User(**user_doc)
    
    # Create session
    session_token = str(uuid.uuid4()) + str(uuid.uuid4())
    expires_at = datetime.now(timezone.utc) + timedelta(days=7)
    
    session = UserSession(
        user_id=user.id,
        session_token=session_token,
        expires_at=expires_at
    )
    
    session_doc = session.model_dump()
    session_doc["expires_at"] = session_doc["expires_at"].isoformat()
    session_doc["created_at"] = session_doc["created_at"].isoformat()
    await db.user_sessions.insert_one(session_doc)
    
    # Set cookie
    response.set_cookie(
        key="session_token",
        value=session_token,
        httponly=True,
        secure=True,
        samesite="none",  # Changed to 'none' for cross-origin requests
        path="/",
        max_age=7 * 24 * 60 * 60
    )
    
    # Remove password from response
    user_dict = user.model_dump()
    user_dict.pop('password', None)
    
    return {"user": user_dict, "message": "Auto-login successful"}

@api_router.get("/auth/me")
async def get_me(request: Request):
    """Get current user"""
    user = await require_auth(request)
    return user.model_dump()

@api_router.post("/auth/logout")
async def logout(request: Request, response: Response):
    """Logout user"""
    session_token = request.cookies.get("session_token")
    if session_token:
        await db.user_sessions.delete_one({"session_token": session_token})
    
    response.delete_cookie(key="session_token", path="/")
    return {"message": "Logged out"}

@api_router.get("/users/{user_id}")
async def get_user(user_id: str, request: Request):
    """Get user by ID"""
    await require_auth(request)
    
    user_doc = await db.users.find_one({"id": user_id})
    if not user_doc:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Remove MongoDB _id and password
    if '_id' in user_doc:
        del user_doc['_id']
    if 'password' in user_doc:
        del user_doc['password']
    
    # Convert datetime
    if isinstance(user_doc.get('created_at'), str):
        user_doc['created_at'] = datetime.fromisoformat(user_doc['created_at'])
    
    return user_doc

@api_router.put("/users/{user_id}/role")
async def update_user_role(user_id: str, role: UserRole, request: Request):
    """Update user role (admin only)"""
    admin = await require_role(request, [UserRole.SYSTEM_ADMIN])
    
    result = await db.users.update_one(
        {"id": user_id},
        {"$set": {"role": role.value}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {"message": "Role updated"}

# ==================== DASHBOARD ENDPOINT ====================

@api_router.get("/dashboard")
async def get_dashboard_stats(request: Request):
    """Get dashboard statistics for all modules"""
    await require_auth(request)
    
    # Vendor Statistics
    all_vendors = await db.vendors.count_documents({})
    active_vendors = await db.vendors.count_documents({"status": VendorStatus.APPROVED.value})
    high_risk_vendors = await db.vendors.count_documents({"risk_category": "high"})
    waiting_due_diligence = await db.vendors.count_documents({"status": VendorStatus.PENDING_DUE_DILIGENCE.value})
    inactive_vendors = await db.vendors.count_documents({"status": VendorStatus.REJECTED.value})
    blacklisted_vendors = await db.vendors.count_documents({"status": VendorStatus.BLACKLISTED.value})
    
    # Tender Statistics
    all_tenders = await db.tenders.count_documents({})
    active_tenders = await db.tenders.count_documents({"status": TenderStatus.PUBLISHED.value})
    
    # Waiting for proposals - published tenders with no proposals or few proposals
    published_tenders = await db.tenders.find({"status": TenderStatus.PUBLISHED.value}).to_list(1000)
    waiting_proposals_count = 0
    waiting_evaluation_count = 0
    
    for tender in published_tenders:
        proposals = await db.proposals.find({"tender_id": tender["id"]}).to_list(1000)
        if len(proposals) == 0:
            waiting_proposals_count += 1
        else:
            # Check if any proposals are not evaluated
            unevaluated = [p for p in proposals if not p.get('evaluation')]
            if len(unevaluated) > 0:
                waiting_evaluation_count += 1
    
    approved_tenders = await db.tenders.count_documents({"status": TenderStatus.AWARDED.value})
    
    # Contract Statistics
    all_contracts = await db.contracts.count_documents({})
    # Active contracts = approved + draft (not expired or pending)
    active_contracts = await db.contracts.count_documents({
        "status": {"$in": [ContractStatus.APPROVED.value, ContractStatus.DRAFT.value]}
    })
    
    # Outsourcing, Cloud, and NOC contracts
    outsourcing_contracts = await db.contracts.count_documents({"outsourcing_classification": "outsourcing"})
    cloud_contracts = await db.contracts.count_documents({"outsourcing_classification": "cloud_computing"})
    noc_contracts = await db.contracts.count_documents({"is_noc": True})
    expired_contracts = await db.contracts.count_documents({"status": ContractStatus.EXPIRED.value})
    
    # Invoice Statistics
    all_invoices = await db.invoices.count_documents({})
    
    # Due invoices - pending or verified status
    due_invoices = await db.invoices.count_documents({
        "status": {"$in": [InvoiceStatus.PENDING.value, InvoiceStatus.VERIFIED.value, InvoiceStatus.APPROVED.value]}
    })
    
    # Purchase Order Statistics
    all_pos = await db.purchase_orders.count_documents({})
    issued_pos = await db.purchase_orders.count_documents({"status": "issued"})
    converted_pos = await db.purchase_orders.count_documents({"status": "converted_to_contract"})
    
    # Calculate total PO value
    pos = await db.purchase_orders.find({}).to_list(1000)
    total_po_value = sum(po.get('total_amount', 0) for po in pos)
    
    # Asset Statistics
    all_assets = await db.assets.count_documents({})
    active_assets = await db.assets.count_documents({"status": "active"})
    under_maintenance_assets = await db.assets.count_documents({"status": "under_maintenance"})
    out_of_service_assets = await db.assets.count_documents({"status": "out_of_service"})
    
    # Warranty statistics
    current_date = datetime.now(timezone.utc)
    in_warranty_assets = 0
    warranty_expiring_assets = 0
    
    assets_with_warranty = await db.assets.find({"warranty_end_date": {"$exists": True, "$ne": None}}, {"_id": 0}).to_list(10000)
    for asset in assets_with_warranty:
        warranty_end = asset.get("warranty_end_date")
        if warranty_end:
            if isinstance(warranty_end, str):
                warranty_end = datetime.fromisoformat(warranty_end)
            if warranty_end > current_date:
                in_warranty_assets += 1
                days_to_expiry = (warranty_end - current_date).days
                if days_to_expiry <= 90:  # Expiring in 90 days
                    warranty_expiring_assets += 1
    
    # OSR Statistics
    all_osr = await db.osr.count_documents({})
    open_osr = await db.osr.count_documents({"status": "open"})
    assigned_osr = await db.osr.count_documents({"status": "assigned"})
    in_progress_osr = await db.osr.count_documents({"status": "in_progress"})
    completed_osr = await db.osr.count_documents({"status": "completed"})
    high_priority_osr = await db.osr.count_documents({"priority": "high"})
    
    return {
        "vendors": {
            "all": all_vendors,
            "active": active_vendors,
            "high_risk": high_risk_vendors,
            "waiting_due_diligence": waiting_due_diligence,
            "inactive": inactive_vendors,
            "blacklisted": blacklisted_vendors
        },
        "tenders": {
            "all": all_tenders,
            "active": active_tenders,
            "waiting_proposals": waiting_proposals_count,
            "waiting_evaluation": waiting_evaluation_count,
            "approved": approved_tenders
        },
        "contracts": {
            "all": all_contracts,
            "active": active_contracts,
            "outsourcing": outsourcing_contracts,
            "cloud": cloud_contracts,
            "noc": noc_contracts,
            "expired": expired_contracts
        },
        "invoices": {
            "all": all_invoices,
            "due": due_invoices
        },
        "resources": {
            "all": await db.resources.count_documents({}),
            "active": await db.resources.count_documents({"status": ResourceStatus.ACTIVE.value}),
            "offshore": await db.resources.count_documents({"work_type": WorkType.OFFSHORE.value}),
            "on_premises": await db.resources.count_documents({"work_type": WorkType.ON_PREMISES.value})
        },
        "purchase_orders": {
            "all": all_pos,
            "issued": issued_pos,
            "converted": converted_pos,
            "total_value": total_po_value
        },
        "assets": {
            "total": all_assets,
            "active": active_assets,
            "under_maintenance": under_maintenance_assets,
            "out_of_service": out_of_service_assets,
            "in_warranty": in_warranty_assets,
            "warranty_expiring": warranty_expiring_assets
        },
        "osr": {
            "total": all_osr,
            "open": open_osr,
            "assigned": assigned_osr,
            "in_progress": in_progress_osr,
            "completed": completed_osr,
            "high_priority": high_priority_osr
        }
    }

# ==================== VENDOR ENDPOINTS ====================
@api_router.post("/vendors")
async def create_vendor(vendor: Vendor, request: Request):
    """Create a new vendor - RBAC: requires create permission"""
    from utils.auth import require_create_permission
    user = await require_create_permission(request, "vendors")
    
    # Calculate detailed risk assessment
    risk_details = {}
    risk_score = 0.0
    
    # Documents check (30 points)
    if not vendor.documents or len(vendor.documents) == 0:
        risk_score += 30
        risk_details["missing_documents"] = {"score": 30, "reason": "No documents uploaded"}
    
    # Bank information check (20 points)
    if not vendor.bank_name or not vendor.iban:
        risk_score += 20
        risk_details["incomplete_banking"] = {"score": 20, "reason": "Missing bank information"}
    
    # CR expiry check (15 points if expiring soon)
    if vendor.cr_expiry_date:
        days_to_expiry = (vendor.cr_expiry_date - datetime.now(timezone.utc)).days
        if days_to_expiry < 90:
            risk_score += 15
            risk_details["cr_expiring_soon"] = {"score": 15, "reason": f"CR expires in {days_to_expiry} days"}
    
    # License check (10 points)
    if not vendor.license_number:
        risk_score += 10
        risk_details["missing_license"] = {"score": 10, "reason": "No license number provided"}
    
    # Number of employees check (10 points if < 5)
    if vendor.number_of_employees < 5:
        risk_score += 10
        risk_details["small_team"] = {"score": 10, "reason": f"Only {vendor.number_of_employees} employees"}
    
    vendor.risk_score = risk_score
    vendor.risk_assessment_details = risk_details
    
    if risk_score >= 50:
        vendor.risk_category = RiskCategory.HIGH
    elif risk_score >= 25:
        vendor.risk_category = RiskCategory.MEDIUM
    else:
        vendor.risk_category = RiskCategory.LOW
    
    # Check if Due Diligence checklist items are provided (from vendor creation)
    checklist_items_present = any([
        vendor.dd_checklist_supporting_documents is not None,
        vendor.dd_checklist_related_party_checked is not None,
        vendor.dd_checklist_sanction_screening is not None,
    ])
    
    # Check if actual DD fields are provided (indicating completed DD during creation)
    dd_fields_present = any([
        vendor.dd_ownership_change_last_year is not None,
        vendor.dd_location_moved_or_closed is not None,
        vendor.dd_bc_rely_on_third_parties is not None,
        vendor.dd_bc_strategy_exists is not None,
        vendor.dd_fraud_internal_last_year is not None,
        vendor.dd_op_criminal_cases_last_3years is not None,
        vendor.dd_hr_background_investigation is not None,
        vendor.dd_safety_procedures_exist is not None
    ])
    
    if dd_fields_present:
        # DD fields provided during creation - mark as completed and approved
        vendor.status = VendorStatus.APPROVED
        vendor.dd_completed = True
        vendor.dd_completed_by = user.id
        vendor.dd_completed_at = datetime.now(timezone.utc)
        vendor.dd_approved_by = user.id
        vendor.dd_approved_at = datetime.now(timezone.utc)
        
        # Recalculate risk score based on DD responses
        vendor_dict = vendor.model_dump()
        dd_adjustment = calculate_dd_risk_adjustment(vendor_dict)
        new_risk_score = max(0, vendor.risk_score + dd_adjustment)
        vendor.risk_score = new_risk_score
        
        # Update risk category based on new score
        if new_risk_score >= 50:
            vendor.risk_category = RiskCategory.HIGH
        elif new_risk_score >= 25:
            vendor.risk_category = RiskCategory.MEDIUM
        else:
            vendor.risk_category = RiskCategory.LOW
    elif checklist_items_present:
        # Only checklist items present - vendor requires DD completion later
        vendor.status = VendorStatus.PENDING_DUE_DILIGENCE
        vendor.dd_completed = False  # Not completed yet, just flagged for DD
    else:
        # No checklist items or DD fields - vendor is approved directly
        vendor.status = VendorStatus.APPROVED
    vendor.created_by = user.id
    
    # Generate vendor number
    vendor.vendor_number = await generate_number("Vendor")
    
    vendor_doc = vendor.model_dump()
    vendor_doc["created_at"] = vendor_doc["created_at"].isoformat()
    vendor_doc["updated_at"] = vendor_doc["updated_at"].isoformat()
    if vendor_doc.get("cr_expiry_date"):
        vendor_doc["cr_expiry_date"] = vendor_doc["cr_expiry_date"].isoformat()
    if vendor_doc.get("license_expiry_date"):
        vendor_doc["license_expiry_date"] = vendor_doc["license_expiry_date"].isoformat()
    
    await db.vendors.insert_one(vendor_doc)
    
    # Create audit log
    audit_log = AuditLog(
        entity_type="vendor",
        entity_id=vendor.id,
        action="created",
        user_id=user.id,
        details=f"Vendor created: {vendor.name_english} (Risk Score: {risk_score}, Status: {vendor.status})"
    )
    audit_doc = audit_log.model_dump()
    audit_doc["timestamp"] = audit_doc["timestamp"].isoformat()
    await db.audit_logs.insert_one(audit_doc)
    
    return vendor.model_dump()

@api_router.get("/vendors")
async def get_vendors(request: Request, status: Optional[VendorStatus] = None, search: Optional[str] = None):
    """Get all vendors with optional search by vendor_number or name - RBAC: requires viewer permission"""
    from utils.auth import require_permission
    from utils.permissions import Permission
    await require_permission(request, "vendors", Permission.VIEWER)
    
    query = {}
    if status:
        query["status"] = status.value
    
    # Add search functionality
    if search:
        query["$or"] = [
            {"vendor_number": {"$regex": search, "$options": "i"}},
            {"name_english": {"$regex": search, "$options": "i"}},
            {"commercial_name": {"$regex": search, "$options": "i"}}
        ]
    
    vendors = await db.vendors.find(query).to_list(1000)
    
    # Convert datetime strings and handle ObjectId
    result = []
    for vendor in vendors:
        # Remove MongoDB _id
        if '_id' in vendor:
            del vendor['_id']
        
        # Convert datetime strings
        if isinstance(vendor.get('created_at'), str):
            vendor['created_at'] = datetime.fromisoformat(vendor['created_at'])
        if isinstance(vendor.get('updated_at'), str):
            vendor['updated_at'] = datetime.fromisoformat(vendor['updated_at'])
        if isinstance(vendor.get('cr_expiry_date'), str):
            vendor['cr_expiry_date'] = datetime.fromisoformat(vendor['cr_expiry_date'])
        if vendor.get('license_expiry_date') and isinstance(vendor.get('license_expiry_date'), str):
            vendor['license_expiry_date'] = datetime.fromisoformat(vendor['license_expiry_date'])
        
        result.append(vendor)
    
    return result

@api_router.get("/vendors/{vendor_id}")
async def get_vendor(vendor_id: str, request: Request):
    """Get vendor by ID - RBAC: requires viewer permission"""
    from utils.auth import require_permission
    from utils.permissions import Permission
    await require_permission(request, "vendors", Permission.VIEWER)
    
    vendor = await db.vendors.find_one({"id": vendor_id})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    
    # Remove MongoDB _id
    if '_id' in vendor:
        del vendor['_id']
    
    # Convert datetime strings
    if isinstance(vendor.get('created_at'), str):
        vendor['created_at'] = datetime.fromisoformat(vendor['created_at'])
    if isinstance(vendor.get('updated_at'), str):
        vendor['updated_at'] = datetime.fromisoformat(vendor['updated_at'])
    if isinstance(vendor.get('cr_expiry_date'), str):
        vendor['cr_expiry_date'] = datetime.fromisoformat(vendor['cr_expiry_date'])
    if vendor.get('license_expiry_date') and isinstance(vendor.get('license_expiry_date'), str):
        vendor['license_expiry_date'] = datetime.fromisoformat(vendor['license_expiry_date'])
    
    return vendor

@api_router.put("/vendors/{vendor_id}")
async def update_vendor(vendor_id: str, vendor_update: Vendor, request: Request):
    """Update vendor information"""
    user = await require_role(request, [UserRole.PROCUREMENT_OFFICER, UserRole.SYSTEM_ADMIN])
    
    # Get existing vendor
    existing_vendor = await db.vendors.find_one({"id": vendor_id})
    if not existing_vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    
    # Recalculate risk assessment
    risk_details = {}
    risk_score = 0.0
    
    if not vendor_update.documents or len(vendor_update.documents) == 0:
        risk_score += 30
        risk_details["missing_documents"] = {"score": 30, "reason": "No documents uploaded"}
    
    if not vendor_update.bank_name or not vendor_update.iban:
        risk_score += 20
        risk_details["incomplete_banking"] = {"score": 20, "reason": "Missing bank information"}
    
    if vendor_update.cr_expiry_date:
        days_to_expiry = (vendor_update.cr_expiry_date - datetime.now(timezone.utc)).days
        if days_to_expiry < 90:
            risk_score += 15
            risk_details["cr_expiring_soon"] = {"score": 15, "reason": f"CR expires in {days_to_expiry} days"}
    
    if not vendor_update.license_number:
        risk_score += 10
        risk_details["missing_license"] = {"score": 10, "reason": "No license number provided"}
    
    if vendor_update.number_of_employees < 5:
        risk_score += 10
        risk_details["small_team"] = {"score": 10, "reason": f"Only {vendor_update.number_of_employees} employees"}
    
    vendor_update.risk_score = risk_score
    vendor_update.risk_assessment_details = risk_details
    
    if risk_score >= 50:
        vendor_update.risk_category = RiskCategory.HIGH
    elif risk_score >= 25:
        vendor_update.risk_category = RiskCategory.MEDIUM
    else:
        vendor_update.risk_category = RiskCategory.LOW
    
    vendor_update.updated_at = datetime.now(timezone.utc)
    vendor_update.id = vendor_id  # Preserve ID
    vendor_update.created_by = existing_vendor.get("created_by")  # Preserve creator
    vendor_update.created_at = datetime.fromisoformat(existing_vendor["created_at"]) if isinstance(existing_vendor.get("created_at"), str) else existing_vendor.get("created_at")
    
    # Track changes
    changes = {}
    for field in ["name_english", "vat_number", "cr_number", "mobile", "email"]:
        old_value = existing_vendor.get(field)
        new_value = getattr(vendor_update, field)
        if old_value != new_value:
            changes[field] = {"old": old_value, "new": new_value}
    
    vendor_doc = vendor_update.model_dump()
    vendor_doc["created_at"] = vendor_doc["created_at"].isoformat()
    vendor_doc["updated_at"] = vendor_doc["updated_at"].isoformat()
    if vendor_doc.get("cr_expiry_date"):
        vendor_doc["cr_expiry_date"] = vendor_doc["cr_expiry_date"].isoformat()
    if vendor_doc.get("license_expiry_date"):
        vendor_doc["license_expiry_date"] = vendor_doc["license_expiry_date"].isoformat()
    
    await db.vendors.update_one({"id": vendor_id}, {"$set": vendor_doc})
    
    # Create audit log
    audit_log = AuditLog(
        entity_type="vendor",
        entity_id=vendor_id,
        action="updated",
        user_id=user.id,
        user_name=user.name,
        changes=changes
    )
    audit_doc = audit_log.model_dump()
    audit_doc["timestamp"] = audit_doc["timestamp"].isoformat()
    await db.audit_logs.insert_one(audit_doc)
    
    return vendor_update.model_dump()

@api_router.get("/vendors/{vendor_id}/audit-log")
async def get_vendor_audit_log(vendor_id: str, request: Request):
    """Get audit log for a vendor"""
    await require_auth(request)
    
    logs = await db.audit_logs.find({"entity_type": "vendor", "entity_id": vendor_id}).sort("timestamp", -1).to_list(100)
    
    # Remove _id and convert timestamps
    result = []
    for log in logs:
        if '_id' in log:
            del log['_id']
        if isinstance(log.get('timestamp'), str):
            log['timestamp'] = datetime.fromisoformat(log['timestamp'])
        result.append(log)
    
    return result

# ==================== DUE DILIGENCE ENDPOINTS ====================

@api_router.put("/vendors/{vendor_id}/due-diligence")
async def update_vendor_due_diligence(vendor_id: str, dd_data: dict, request: Request):
    """Update vendor due diligence questionnaire - Auto-approves and recalculates risk"""
    user = await require_role(request, [UserRole.PROCUREMENT_OFFICER, UserRole.SYSTEM_ADMIN])
    
    vendor = await db.vendors.find_one({"id": vendor_id})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    
    # Add all dd_ fields from dd_data
    for key, value in dd_data.items():
        if key.startswith('dd_'):
            vendor[key] = value
    
    # Recalculate risk score based on DD responses
    base_risk = vendor.get('risk_score', 0.0)
    dd_adjustment = calculate_dd_risk_adjustment(vendor)
    new_risk_score = max(0, base_risk + dd_adjustment)
    
    # Determine new risk category
    if new_risk_score >= 50:
        new_risk_category = RiskCategory.HIGH.value
    elif new_risk_score >= 25:
        new_risk_category = RiskCategory.MEDIUM.value
    else:
        new_risk_category = RiskCategory.LOW.value
    
    # Update all dd_ prefixed fields and auto-approve
    update_fields = {
        "dd_completed": True,
        "dd_completed_by": user.id,
        "dd_completed_at": datetime.now(timezone.utc).isoformat(),
        "dd_approved_by": user.id,
        "dd_approved_at": datetime.now(timezone.utc).isoformat(),
        "status": VendorStatus.APPROVED.value,
        "risk_score": new_risk_score,
        "risk_category": new_risk_category,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Add all dd_ fields from dd_data
    for key, value in dd_data.items():
        if key.startswith('dd_'):
            update_fields[key] = value
    
    await db.vendors.update_one(
        {"id": vendor_id},
        {"$set": update_fields}
    )
    
    # Auto-approve all pending contracts for this vendor
    await db.contracts.update_many(
        {
            "vendor_id": vendor_id,
            "status": ContractStatus.PENDING_DUE_DILIGENCE.value
        },
        {"$set": {
            "status": ContractStatus.APPROVED.value,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {
        "message": "Due diligence completed and auto-approved. Vendor and contracts status updated.",
        "new_risk_score": new_risk_score,
        "new_risk_category": new_risk_category
    }

@api_router.post("/vendors/{vendor_id}/due-diligence/approve")
async def approve_vendor_due_diligence(vendor_id: str, request: Request):
    """Approve vendor due diligence and change status back to approved"""
    user = await require_role(request, [UserRole.PROCUREMENT_OFFICER, UserRole.SYSTEM_ADMIN])
    
    vendor = await db.vendors.find_one({"id": vendor_id})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    
    if not vendor.get('dd_completed'):
        raise HTTPException(status_code=400, detail="Due diligence not completed yet")
    
    # Update vendor status to approved and set approval info
    await db.vendors.update_one(
        {"id": vendor_id},
        {"$set": {
            "status": VendorStatus.APPROVED.value,
            "dd_approved_by": user.id,
            "dd_approved_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Update all contracts for this vendor from pending_due_diligence to approved
    await db.contracts.update_many(
        {
            "vendor_id": vendor_id,
            "status": ContractStatus.PENDING_DUE_DILIGENCE.value
        },
        {"$set": {
            "status": ContractStatus.APPROVED.value,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "Due diligence approved successfully. Vendor and contracts status updated."}

@api_router.post("/vendors/{vendor_id}/blacklist")
async def blacklist_vendor(vendor_id: str, request: Request):
    """Blacklist a vendor - PD Officer or Admin only"""
    user = await require_role(request, [UserRole.PD_OFFICER, UserRole.ADMIN, UserRole.PROCUREMENT_OFFICER, UserRole.SYSTEM_ADMIN])
    
    vendor = await db.vendors.find_one({"id": vendor_id})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    
    # Update vendor status to blacklisted
    await db.vendors.update_one(
        {"id": vendor_id},
        {"$set": {
            "status": VendorStatus.BLACKLISTED.value,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Terminate all active contracts for this vendor
    await db.contracts.update_many(
        {
            "vendor_id": vendor_id,
            "status": {"$in": [ContractStatus.ACTIVE.value, ContractStatus.APPROVED.value]}
        },
        {"$set": {
            "terminated": True,
            "terminated_by": user.id,
            "terminated_at": datetime.now(timezone.utc).isoformat(),
            "termination_reason": "Vendor blacklisted",
            "status": ContractStatus.EXPIRED.value,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "Vendor blacklisted and all active contracts terminated"}

# ==================== TENDER ENDPOINTS ====================
@api_router.post("/tenders")
async def create_tender(tender: Tender, request: Request):
    """Create new tender - Auto-approved with generated number"""
    user = await require_role(request, [UserRole.PROCUREMENT_OFFICER])
    
    tender.created_by = user.id
    
    # Auto-approve and generate tender number
    tender.status = TenderStatus.PUBLISHED
    tender.tender_number = await generate_number("Tender")
    
    tender_doc = tender.model_dump()
    tender_doc["deadline"] = tender_doc["deadline"].isoformat()
    tender_doc["created_at"] = tender_doc["created_at"].isoformat()
    tender_doc["updated_at"] = tender_doc["updated_at"].isoformat()
    
    await db.tenders.insert_one(tender_doc)
    
    # Return without MongoDB _id
    result = tender.model_dump()
    return result

@api_router.get("/tenders")
async def get_tenders(request: Request, status: Optional[TenderStatus] = None, search: Optional[str] = None):
    """Get all tenders with optional search by tender_number or title"""
    user = await require_role(request, [UserRole.PROCUREMENT_OFFICER, UserRole.PROJECT_MANAGER, UserRole.SYSTEM_ADMIN])
    
    query = {}
    if status:
        query["status"] = status.value
    
    # Add search functionality
    if search:
        query["$or"] = [
            {"tender_number": {"$regex": search, "$options": "i"}},
            {"title": {"$regex": search, "$options": "i"}},
            {"project_name": {"$regex": search, "$options": "i"}}
        ]
    
    tenders = await db.tenders.find(query).to_list(1000)
    
    result = []
    for tender in tenders:
        # Remove MongoDB _id
        if '_id' in tender:
            del tender['_id']
        
        if isinstance(tender.get('deadline'), str):
            tender['deadline'] = datetime.fromisoformat(tender['deadline'])
        if isinstance(tender.get('created_at'), str):
            tender['created_at'] = datetime.fromisoformat(tender['created_at'])
        if isinstance(tender.get('updated_at'), str):
            tender['updated_at'] = datetime.fromisoformat(tender['updated_at'])
        
        result.append(tender)
    
    return result

@api_router.get("/tenders/{tender_id}")
async def get_tender(tender_id: str, request: Request):
    """Get tender by ID"""
    await require_auth(request)
    
    tender = await db.tenders.find_one({"id": tender_id})
    if not tender:
        raise HTTPException(status_code=404, detail="Tender not found")
    
    # Remove MongoDB _id
    if '_id' in tender:
        del tender['_id']
    
    if isinstance(tender.get('deadline'), str):
        tender['deadline'] = datetime.fromisoformat(tender['deadline'])
    if isinstance(tender.get('created_at'), str):
        tender['created_at'] = datetime.fromisoformat(tender['created_at'])
    if isinstance(tender.get('updated_at'), str):
        tender['updated_at'] = datetime.fromisoformat(tender['updated_at'])
    
    return tender

@api_router.get("/tenders/approved/list")
async def get_approved_tenders(request: Request):
    """Get list of approved tenders for contract creation"""
    await require_auth(request)
    
    tenders = await db.tenders.find({"status": TenderStatus.PUBLISHED.value}).to_list(1000)
    
    result = []
    for tender in tenders:
        # Remove MongoDB _id
        if '_id' in tender:
            del tender['_id']
        
        # Only return essential fields for dropdown
        result.append({
            "id": tender.get("id"),
            "tender_number": tender.get("tender_number"),
            "title": tender.get("title"),
            "project_name": tender.get("project_name"),
            "requirements": tender.get("requirements"),
            "budget": tender.get("budget")
        })
    
    return result

@api_router.put("/tenders/{tender_id}")
async def update_tender(tender_id: str, tender: Tender, request: Request):
    """Update tender"""
    user = await require_role(request, [UserRole.PROCUREMENT_OFFICER])
    
    # Check if tender exists
    existing_tender = await db.tenders.find_one({"id": tender_id})
    if not existing_tender:
        raise HTTPException(status_code=404, detail="Tender not found")
    
    # Prepare update data
    update_data = {
        "title": tender.title,
        "description": tender.description,
        "project_name": tender.project_name,
        "requirements": tender.requirements,
        "budget": tender.budget,
        "deadline": tender.deadline.isoformat(),
        "invited_vendors": tender.invited_vendors,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Update tender
    result = await db.tenders.update_one(
        {"id": tender_id},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Tender not found")
    
    return {"message": "Tender updated successfully"}

@api_router.put("/tenders/{tender_id}/publish")
async def publish_tender(tender_id: str, request: Request):
    """Publish tender"""
    await require_role(request, [UserRole.PROCUREMENT_OFFICER])
    
    result = await db.tenders.update_one(
        {"id": tender_id},
        {
            "$set": {
                "status": TenderStatus.PUBLISHED.value,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
        }
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Tender not found")
    
    return {"message": "Tender published"}

@api_router.post("/tenders/{tender_id}/proposals")
async def submit_proposal(tender_id: str, proposal: Proposal, request: Request):
    """Submit proposal for tender (Procurement Officer can submit on behalf of vendor)"""
    await require_role(request, [UserRole.PROCUREMENT_OFFICER, UserRole.SYSTEM_ADMIN])
    
    proposal.tender_id = tender_id
    
    # Generate proposal number
    proposal.proposal_number = await generate_number("Proposal")
    
    # vendor_id should be provided in the proposal object
    proposal_doc = proposal.model_dump()
    proposal_doc["submitted_at"] = proposal_doc["submitted_at"].isoformat()
    
    await db.proposals.insert_one(proposal_doc)
    
    # Return without MongoDB _id
    result = proposal.model_dump()
    return result

@api_router.get("/tenders/{tender_id}/proposals")
async def get_tender_proposals(tender_id: str, request: Request):
    """Get all proposals for a tender"""
    await require_role(request, [UserRole.PROCUREMENT_OFFICER, UserRole.PROJECT_MANAGER])
    
    proposals = await db.proposals.find({"tender_id": tender_id}).to_list(1000)
    
    result = []
    for proposal in proposals:
        # Remove MongoDB _id
        if '_id' in proposal:
            del proposal['_id']
        
        if isinstance(proposal.get('submitted_at'), str):
            proposal['submitted_at'] = datetime.fromisoformat(proposal['submitted_at'])
        
        result.append(proposal)
    
    return result

class ProposalEvaluationRequest(BaseModel):
    proposal_id: str
    vendor_reliability_stability: float = Field(ge=1, le=5)
    delivery_warranty_backup: float = Field(ge=1, le=5)
    technical_experience: float = Field(ge=1, le=5)
    cost_score: float = Field(ge=1, le=5)
    meets_requirements: float = Field(ge=1, le=5)

@api_router.post("/tenders/{tender_id}/proposals/{proposal_id}/evaluate")
async def evaluate_proposal(tender_id: str, proposal_id: str, evaluation: ProposalEvaluationRequest, request: Request):
    """Evaluate a single proposal with detailed criteria"""
    user = await require_role(request, [UserRole.PROCUREMENT_OFFICER, UserRole.PROJECT_MANAGER])
    
    # Find proposal
    proposal = await db.proposals.find_one({"id": proposal_id, "tender_id": tender_id})
    if not proposal:
        raise HTTPException(status_code=404, detail="Proposal not found")
    
    # Calculate weighted scores (weights: 20%, 20%, 10%, 10%, 40% = 100% total)
    vendor_reliability_weighted = evaluation.vendor_reliability_stability * 0.20
    delivery_warranty_weighted = evaluation.delivery_warranty_backup * 0.20
    technical_experience_weighted = evaluation.technical_experience * 0.10
    cost_weighted = evaluation.cost_score * 0.10
    meets_requirements_weighted = evaluation.meets_requirements * 0.40
    
    total_score = (
        vendor_reliability_weighted + 
        delivery_warranty_weighted + 
        technical_experience_weighted + 
        cost_weighted +
        meets_requirements_weighted
    )
    
    # Create evaluation object
    evaluation_data = {
        "vendor_reliability_stability": evaluation.vendor_reliability_stability,
        "delivery_warranty_backup": evaluation.delivery_warranty_backup,
        "technical_experience": evaluation.technical_experience,
        "cost_score": evaluation.cost_score,
        "meets_requirements": evaluation.meets_requirements,
        "vendor_reliability_weighted": vendor_reliability_weighted,
        "delivery_warranty_weighted": delivery_warranty_weighted,
        "technical_experience_weighted": technical_experience_weighted,
        "cost_weighted": cost_weighted,
        "meets_requirements_weighted": meets_requirements_weighted,
        "total_score": total_score
    }
    
    # Update proposal with evaluation
    await db.proposals.update_one(
        {"id": proposal_id},
        {
            "$set": {
                "evaluation": evaluation_data,
                "evaluated_by": user.id,
                "evaluated_at": datetime.now(timezone.utc).isoformat(),
                "final_score": total_score
            }
        }
    )
    
    return {
        "message": "Proposal evaluated successfully",
        "evaluation": evaluation_data,
        "total_score": total_score
    }

@api_router.post("/tenders/{tender_id}/evaluate")
async def evaluate_all_proposals(tender_id: str, request: Request):
    """Get evaluation summary for all proposals in a tender"""
    await require_role(request, [UserRole.PROCUREMENT_OFFICER, UserRole.PROJECT_MANAGER])
    
    proposals = await db.proposals.find({"tender_id": tender_id}).to_list(1000)
    
    if not proposals:
        return {"message": "No proposals to evaluate", "proposals": []}
    
    # Calculate cost scores automatically based on lowest price
    min_price = min(p["financial_proposal"] for p in proposals if p.get("financial_proposal"))
    
    evaluated_proposals = []
    for proposal in proposals:
        # Auto-calculate cost score (lowest price gets 5, others scaled)
        if min_price > 0:
            cost_score = (min_price / proposal["financial_proposal"]) * 5
        else:
            cost_score = 3.0  # Default if no prices
        
        # Get vendor name
        vendor = await db.vendors.find_one({"id": proposal["vendor_id"]})
        vendor_name = vendor.get("name_english", vendor.get("company_name", "Unknown")) if vendor else "Unknown"
        
        evaluated_proposals.append({
            "proposal_id": proposal["id"],
            "vendor_id": proposal["vendor_id"],
            "vendor_name": vendor_name,
            "financial_proposal": proposal["financial_proposal"],
            "suggested_cost_score": round(cost_score, 2),
            "evaluation": proposal.get("evaluation"),
            "final_score": proposal.get("final_score", 0.0),
            "evaluated": proposal.get("evaluation") is not None
        })
    
    # Sort by final score descending
    evaluated_proposals.sort(key=lambda x: x["final_score"], reverse=True)
    
    return {
        "tender_id": tender_id,
        "total_proposals": len(proposals),
        "evaluated_count": sum(1 for p in proposals if p.get("evaluation")),
        "proposals": evaluated_proposals
    }

@api_router.post("/tenders/{tender_id}/award")
async def award_tender(tender_id: str, vendor_id: str, request: Request):
    """Award tender to vendor"""
    user = await require_role(request, [UserRole.PROJECT_MANAGER])
    
    # Update tender
    result = await db.tenders.update_one(
        {"id": tender_id},
        {
            "$set": {
                "status": TenderStatus.AWARDED.value,
                "awarded_to": vendor_id,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
        }
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Tender not found")
    
    return {"message": "Tender awarded", "vendor_id": vendor_id}

# ==================== CONTRACT ENDPOINTS ====================
@api_router.post("/contracts")
async def create_contract(contract: Contract, request: Request):
    """Create new contract - Auto-approved with generated number"""
    user = await require_role(request, [UserRole.PROCUREMENT_OFFICER])
    
    # Verify tender exists
    tender = await db.tenders.find_one({"id": contract.tender_id})
    if not tender:
        raise HTTPException(status_code=404, detail="Tender not found")
    
    # Verify vendor exists
    vendor = await db.vendors.find_one({"id": contract.vendor_id})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    
    contract.created_by = user.id
    
    # Generate contract number
    contract.contract_number = await generate_number("Contract")
    
    # Calculate outsourcing classification based on questionnaire
    contract_doc = contract.model_dump()
    contract.outsourcing_classification = determine_outsourcing_classification(contract_doc)
    contract_doc["outsourcing_classification"] = contract.outsourcing_classification
    
    # Determine NOC requirement
    vendor_type = vendor.get('vendor_type', 'local')
    contract.is_noc = determine_noc_requirement(contract_doc, vendor_type)
    contract_doc["is_noc"] = contract.is_noc
    
    # Check if Due Diligence is required
    vendor_risk = vendor.get('risk_category', 'low')
    vendor_status = vendor.get('status', 'approved')
    classification = contract.outsourcing_classification
    vendor_dd_completed = vendor.get('dd_completed', False)
    
    requires_due_diligence = (
        vendor_risk == 'high' or 
        classification == 'outsourcing' or 
        classification == 'cloud_computing'
    )
    
    # Check if vendor DD is pending or not completed
    vendor_dd_pending = (
        vendor_status == VendorStatus.PENDING_DUE_DILIGENCE.value or 
        (requires_due_diligence and not vendor_dd_completed)
    )
    
    if vendor_dd_pending:
        # Set contract to pending_due_diligence status
        contract.status = ContractStatus.PENDING_DUE_DILIGENCE
        contract_doc["status"] = ContractStatus.PENDING_DUE_DILIGENCE.value
        
        # Update vendor to require and pending due diligence (if not already set)
        if vendor_status != VendorStatus.PENDING_DUE_DILIGENCE.value:
            await db.vendors.update_one(
                {"id": contract.vendor_id},
                {"$set": {
                    "status": VendorStatus.PENDING_DUE_DILIGENCE.value,
                    "dd_required": True
                }}
            )
    else:
        # Auto-approve if no due diligence required or already completed
        contract.status = ContractStatus.APPROVED
        contract_doc["status"] = ContractStatus.APPROVED.value
    
    contract_doc["start_date"] = contract_doc["start_date"].isoformat()
    contract_doc["end_date"] = contract_doc["end_date"].isoformat()
    contract_doc["created_at"] = contract_doc["created_at"].isoformat()
    contract_doc["updated_at"] = contract_doc["updated_at"].isoformat()
    
    await db.contracts.insert_one(contract_doc)
    
    # Return without MongoDB _id
    result = contract.model_dump()
    return result

@api_router.get("/contracts")
async def get_contracts(request: Request, status: Optional[ContractStatus] = None, search: Optional[str] = None):
    """Get all contracts with optional search by contract_number or title"""
    await require_role(request, [UserRole.PROCUREMENT_OFFICER, UserRole.PROJECT_MANAGER, UserRole.SYSTEM_ADMIN])
    
    query = {}
    if status:
        query["status"] = status.value
    
    # Add search functionality
    if search:
        query["$or"] = [
            {"contract_number": {"$regex": search, "$options": "i"}},
            {"title": {"$regex": search, "$options": "i"}}
        ]
    
    contracts = await db.contracts.find(query).to_list(1000)
    
    result = []
    now = datetime.now(timezone.utc)
    
    for contract in contracts:
        # Remove MongoDB _id
        if '_id' in contract:
            del contract['_id']
        
        if isinstance(contract.get('start_date'), str):
            contract['start_date'] = datetime.fromisoformat(contract['start_date'])
        if isinstance(contract.get('end_date'), str):
            contract['end_date'] = datetime.fromisoformat(contract['end_date'])
        if isinstance(contract.get('created_at'), str):
            contract['created_at'] = datetime.fromisoformat(contract['created_at'])
        if isinstance(contract.get('updated_at'), str):
            contract['updated_at'] = datetime.fromisoformat(contract['updated_at'])
        
        # Auto-mark expired contracts
        if (contract.get('end_date') and 
            contract['end_date'] < now and 
            contract.get('status') not in [ContractStatus.EXPIRED.value] and
            not contract.get('terminated')):
            await db.contracts.update_one(
                {"id": contract['id']},
                {"$set": {
                    "status": ContractStatus.EXPIRED.value,
                    "updated_at": now.isoformat()
                }}
            )
            contract['status'] = ContractStatus.EXPIRED.value
        
        result.append(contract)
    
    return result

@api_router.get("/contracts/{contract_id}")
async def get_contract(contract_id: str, request: Request):
    """Get contract by ID"""
    await require_auth(request)
    
    contract = await db.contracts.find_one({"id": contract_id})
    if not contract:
        raise HTTPException(status_code=404, detail="Contract not found")
    
    # Remove MongoDB _id
    if '_id' in contract:
        del contract['_id']
    
    if isinstance(contract.get('start_date'), str):
        contract['start_date'] = datetime.fromisoformat(contract['start_date'])
    if isinstance(contract.get('end_date'), str):
        contract['end_date'] = datetime.fromisoformat(contract['end_date'])
    if isinstance(contract.get('created_at'), str):
        contract['created_at'] = datetime.fromisoformat(contract['created_at'])
    if isinstance(contract.get('updated_at'), str):
        contract['updated_at'] = datetime.fromisoformat(contract['updated_at'])
    
    return contract

@api_router.put("/contracts/{contract_id}/approve")
async def approve_contract(contract_id: str, request: Request):
    """Approve contract"""
    user = await require_role(request, [UserRole.PROJECT_MANAGER])
    
    result = await db.contracts.update_one(
        {"id": contract_id},
        {
            "$set": {
                "status": ContractStatus.APPROVED.value,
                "approved_by": user.id,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
        }
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Contract not found")
    
    return {"message": "Contract approved"}

@api_router.post("/contracts/{contract_id}/terminate")
async def terminate_contract(contract_id: str, request: Request, reason: str = "Manual termination"):
    """Terminate a contract - PD Officer or Admin only"""
    user = await require_role(request, [UserRole.PD_OFFICER, UserRole.ADMIN, UserRole.PROCUREMENT_OFFICER, UserRole.SYSTEM_ADMIN])
    
    contract = await db.contracts.find_one({"id": contract_id})
    if not contract:
        raise HTTPException(status_code=404, detail="Contract not found")
    
    # Update contract to terminated and expired
    result = await db.contracts.update_one(
        {"id": contract_id},
        {"$set": {
            "terminated": True,
            "terminated_by": user.id,
            "terminated_at": datetime.now(timezone.utc).isoformat(),
            "termination_reason": reason,
            "status": ContractStatus.EXPIRED.value,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "Contract terminated successfully"}

@api_router.put("/contracts/{contract_id}")
async def update_contract(contract_id: str, contract_data: dict, request: Request):
    """Update contract details"""
    user = await require_role(request, [UserRole.PROCUREMENT_OFFICER, UserRole.SYSTEM_ADMIN, UserRole.PD_OFFICER, UserRole.ADMIN])
    
    contract = await db.contracts.find_one({"id": contract_id})
    if not contract:
        raise HTTPException(status_code=404, detail="Contract not found")
    
    # Only allow updating certain fields
    allowed_fields = ['title', 'sow', 'sla', 'milestones', 'value', 'start_date', 'end_date']
    update_data = {k: v for k, v in contract_data.items() if k in allowed_fields}
    
    # Convert dates to ISO format if they're provided
    if 'start_date' in update_data and update_data['start_date']:
        if not isinstance(update_data['start_date'], str):
            update_data['start_date'] = update_data['start_date'].isoformat()
    if 'end_date' in update_data and update_data['end_date']:
        if not isinstance(update_data['end_date'], str):
            update_data['end_date'] = update_data['end_date'].isoformat()
    
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    await db.contracts.update_one(
        {"id": contract_id},
        {"$set": update_data}
    )
    
    return {"message": "Contract updated successfully"}

@api_router.get("/contracts/expiring")
async def get_expiring_contracts(request: Request, days: int = 30):
    """Get contracts expiring soon"""
    await require_role(request, [UserRole.PROCUREMENT_OFFICER, UserRole.PROJECT_MANAGER])
    
    expiry_date = datetime.now(timezone.utc) + timedelta(days=days)
    
    contracts = await db.contracts.find({
        "end_date": {
            "$gte": datetime.now(timezone.utc).isoformat(),
            "$lte": expiry_date.isoformat()
        },
        "status": ContractStatus.ACTIVE.value
    }).to_list(1000)
    
    for contract in contracts:
        if isinstance(contract.get('start_date'), str):
            contract['start_date'] = datetime.fromisoformat(contract['start_date'])
        if isinstance(contract.get('end_date'), str):
            contract['end_date'] = datetime.fromisoformat(contract['end_date'])
        if isinstance(contract.get('created_at'), str):
            contract['created_at'] = datetime.fromisoformat(contract['created_at'])
        if isinstance(contract.get('updated_at'), str):
            contract['updated_at'] = datetime.fromisoformat(contract['updated_at'])
    
    return contracts

# ==================== INVOICE ENDPOINTS ====================
@api_router.post("/invoices")
async def submit_invoice(invoice: Invoice, request: Request):
    """Submit invoice with duplicate validation"""
    await require_role(request, [UserRole.PROCUREMENT_OFFICER, UserRole.SYSTEM_ADMIN])
    
    # Verify contract exists
    contract = await db.contracts.find_one({"id": invoice.contract_id})
    if not contract:
        raise HTTPException(status_code=404, detail="Contract not found")
    
    # Check for duplicate invoice (same invoice_number and vendor_id)
    if invoice.invoice_number:
        existing_invoice = await db.invoices.find_one({
            "invoice_number": invoice.invoice_number,
            "vendor_id": invoice.vendor_id
        })
        if existing_invoice:
            raise HTTPException(
                status_code=400, 
                detail=f"Duplicate invoice detected! Invoice number '{invoice.invoice_number}' already exists for this vendor."
            )
    else:
        # Auto-generate invoice number if not provided
        invoice.invoice_number = await generate_number("Invoice")
    
    # Auto-approve
    invoice.status = InvoiceStatus.APPROVED
    
    # vendor_id should be provided in the invoice object
    invoice_doc = invoice.model_dump()
    invoice_doc["submitted_at"] = invoice_doc["submitted_at"].isoformat()
    if invoice_doc.get("verified_at"):
        invoice_doc["verified_at"] = invoice_doc["verified_at"].isoformat()
    if invoice_doc.get("approved_at"):
        invoice_doc["approved_at"] = invoice_doc["approved_at"].isoformat()
    if invoice_doc.get("paid_at"):
        invoice_doc["paid_at"] = invoice_doc["paid_at"].isoformat()
    
    await db.invoices.insert_one(invoice_doc)
    
    # Notify procurement officers
    vendor = await db.vendors.find_one({"id": invoice.vendor_id})
    vendor_name = vendor.get("name_english", "Unknown") if vendor else "Unknown"
    
    procurement_users = await db.users.find({"role": UserRole.PROCUREMENT_OFFICER.value}).to_list(100)
    for po_user in procurement_users:
        notif = Notification(
            user_id=po_user["id"],
            title="New Invoice Submitted",
            message=f"Invoice {invoice.invoice_number} submitted for vendor {vendor_name}",
            type="approval"
        )
        notif_doc = notif.model_dump()
        notif_doc["created_at"] = notif_doc["created_at"].isoformat()
        await db.notifications.insert_one(notif_doc)
    
    return invoice.model_dump()

@api_router.get("/invoices")
async def get_invoices(request: Request, status: Optional[InvoiceStatus] = None, search: Optional[str] = None):
    """Get all invoices with optional search by invoice_number"""
    await require_role(request, [UserRole.PROCUREMENT_OFFICER, UserRole.PROJECT_MANAGER, UserRole.SYSTEM_ADMIN])
    
    query = {}
    if status:
        query["status"] = status.value
    
    # Add search functionality
    if search:
        query["$or"] = [
            {"invoice_number": {"$regex": search, "$options": "i"}},
            {"description": {"$regex": search, "$options": "i"}}
        ]
    
    invoices = await db.invoices.find(query).to_list(1000)
    
    result = []
    for invoice in invoices:
        # Remove MongoDB _id
        if '_id' in invoice:
            del invoice['_id']
        
        if isinstance(invoice.get('submitted_at'), str):
            invoice['submitted_at'] = datetime.fromisoformat(invoice['submitted_at'])
        if invoice.get('verified_at') and isinstance(invoice['verified_at'], str):
            invoice['verified_at'] = datetime.fromisoformat(invoice['verified_at'])
        if invoice.get('approved_at') and isinstance(invoice['approved_at'], str):
            invoice['approved_at'] = datetime.fromisoformat(invoice['approved_at'])
        if invoice.get('paid_at') and isinstance(invoice['paid_at'], str):
            invoice['paid_at'] = datetime.fromisoformat(invoice['paid_at'])
        
        result.append(invoice)
    
    return result

@api_router.get("/invoices/{invoice_id}")
async def get_invoice(invoice_id: str, request: Request):
    """Get invoice by ID"""
    await require_auth(request)
    
    invoice = await db.invoices.find_one({"id": invoice_id})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    # Remove MongoDB _id
    if '_id' in invoice:
        del invoice['_id']
    
    if isinstance(invoice.get('submitted_at'), str):
        invoice['submitted_at'] = datetime.fromisoformat(invoice['submitted_at'])
    if invoice.get('verified_at') and isinstance(invoice['verified_at'], str):
        invoice['verified_at'] = datetime.fromisoformat(invoice['verified_at'])
    if invoice.get('approved_at') and isinstance(invoice['approved_at'], str):
        invoice['approved_at'] = datetime.fromisoformat(invoice['approved_at'])
    if invoice.get('paid_at') and isinstance(invoice['paid_at'], str):
        invoice['paid_at'] = datetime.fromisoformat(invoice['paid_at'])
    
    return invoice

@api_router.put("/invoices/{invoice_id}")
async def update_invoice(invoice_id: str, invoice_data: dict, request: Request):
    """Update invoice details"""
    user = await require_role(request, [UserRole.PROCUREMENT_OFFICER, UserRole.SYSTEM_ADMIN])
    
    invoice = await db.invoices.find_one({"id": invoice_id})
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    # Only allow updating certain fields
    allowed_fields = ['amount', 'description', 'milestone_reference', 'documents']
    update_data = {k: v for k, v in invoice_data.items() if k in allowed_fields}
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    await db.invoices.update_one(
        {"id": invoice_id},
        {"$set": update_data}
    )
    
    return {"message": "Invoice updated successfully"}

@api_router.put("/invoices/{invoice_id}/verify")
async def verify_invoice(invoice_id: str, request: Request):
    """Verify invoice (Procurement Officer)"""
    user = await require_role(request, [UserRole.PROCUREMENT_OFFICER])
    
    result = await db.invoices.update_one(
        {"id": invoice_id},
        {
            "$set": {
                "status": InvoiceStatus.VERIFIED.value,
                "verified_by": user.id,
                "verified_at": datetime.now(timezone.utc).isoformat()
            }
        }
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    return {"message": "Invoice verified"}

@api_router.put("/invoices/{invoice_id}/approve")
async def approve_invoice(invoice_id: str, request: Request):
    """Approve invoice (Project Manager)"""
    user = await require_role(request, [UserRole.PROJECT_MANAGER])
    
    result = await db.invoices.update_one(
        {"id": invoice_id},
        {
            "$set": {
                "status": InvoiceStatus.APPROVED.value,
                "approved_by": user.id,
                "approved_at": datetime.now(timezone.utc).isoformat()
            }
        }
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    return {"message": "Invoice approved"}

# ==================== PURCHASE ORDER ENDPOINTS ====================
@api_router.post("/purchase-orders")
async def create_purchase_order(po: PurchaseOrder, request: Request):
    """Create new purchase order with risk assessment"""
    user = await require_role(request, [UserRole.PROCUREMENT_OFFICER, UserRole.SYSTEM_ADMIN, UserRole.PD_OFFICER, UserRole.ADMIN, UserRole.REQUESTER])
    
    # Generate PO number
    year = datetime.now(timezone.utc).strftime('%y')
    count = await db.purchase_orders.count_documents({}) + 1
    po.po_number = f"PO-{year}-{count:04d}"
    
    # Calculate total amount
    po.total_amount = sum(item.total for item in po.items)
    
    # Check if amount > 1,000,000 SAR
    po.amount_over_million = po.total_amount > 1000000
    
    # Determine if contract is required (any yes answer or amount > 1M)
    po.requires_contract = (
        po.has_data_access or 
        po.has_onsite_presence or 
        po.has_implementation or 
        po.duration_more_than_year or 
        po.amount_over_million
    )
    
    # If no contract required, automatically issue the PO
    if not po.requires_contract:
        po.status = POStatus.ISSUED
    
    po.created_by = user.id
    po_dict = po.model_dump()
    
    await db.purchase_orders.insert_one(po_dict)
    
    return {
        "message": "Purchase order created successfully",
        "po_number": po.po_number,
        "requires_contract": po.requires_contract,
        "status": po.status
    }

@api_router.get("/purchase-orders")
async def get_purchase_orders(request: Request):
    """Get all purchase orders"""
    await require_role(request, [UserRole.PROCUREMENT_OFFICER, UserRole.SYSTEM_ADMIN, UserRole.PD_OFFICER, UserRole.ADMIN])
    
    pos = await db.purchase_orders.find({}, {"_id": 0}).to_list(1000)
    return pos

@api_router.get("/purchase-orders/{po_id}")
async def get_purchase_order(po_id: str, request: Request):
    """Get purchase order by ID"""
    await require_role(request, [UserRole.PROCUREMENT_OFFICER, UserRole.SYSTEM_ADMIN, UserRole.PD_OFFICER, UserRole.ADMIN])
    
    po = await db.purchase_orders.find_one({"id": po_id}, {"_id": 0})
    if not po:
        raise HTTPException(status_code=404, detail="Purchase order not found")
    
    return po

@api_router.post("/purchase-orders/{po_id}/convert-to-contract")
async def convert_po_to_contract(po_id: str, contract_data: dict, request: Request):
    """Convert PO to contract"""
    user = await require_role(request, [UserRole.PROCUREMENT_OFFICER, UserRole.SYSTEM_ADMIN, UserRole.PD_OFFICER, UserRole.ADMIN])
    
    po = await db.purchase_orders.find_one({"id": po_id})
    if not po:
        raise HTTPException(status_code=404, detail="Purchase order not found")
    
    # Create contract from PO data
    contract = Contract(
        tender_id=po.get('tender_id'),
        vendor_id=po['vendor_id'],
        title=contract_data.get('title', f"Contract for PO {po['po_number']}"),
        sow=contract_data.get('sow', 'Contract created from Purchase Order'),
        sla=contract_data.get('sla', 'Standard SLA'),
        value=po['total_amount'],
        start_date=contract_data.get('start_date'),
        end_date=contract_data.get('end_date'),
        created_by=user.id,
        status=ContractStatus.DRAFT
    )
    
    # Generate contract number
    year = datetime.now(timezone.utc).strftime('%y')
    count = await db.contracts.count_documents({}) + 1
    contract.contract_number = f"CNT-{year}-{count:04d}"
    
    contract_dict = contract.model_dump()
    await db.contracts.insert_one(contract_dict)
    
    # Update PO status
    await db.purchase_orders.update_one(
        {"id": po_id},
        {"$set": {
            "status": POStatus.CONVERTED_TO_CONTRACT.value,
            "converted_to_contract": True,
            "contract_id": contract.id,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {
        "message": "Purchase order converted to contract successfully",
        "contract_id": contract.id,
        "contract_number": contract.contract_number
    }

# ==================== RESOURCE ENDPOINTS ====================
@api_router.post("/resources")
async def create_resource(resource: Resource, request: Request):
    """Create new resource based on approved contract and vendor"""
    user = await require_role(request, [UserRole.PROCUREMENT_OFFICER, UserRole.SYSTEM_ADMIN, UserRole.PD_OFFICER, UserRole.ADMIN, UserRole.REQUESTER])
    
    # Verify contract exists and is approved
    contract = await db.contracts.find_one({"id": resource.contract_id})
    if not contract:
        raise HTTPException(status_code=404, detail="Contract not found")
    
    if contract.get('status') not in ['active', 'approved']:
        raise HTTPException(status_code=400, detail="Contract must be active or approved")
    
    # Verify vendor exists and is approved
    vendor = await db.vendors.find_one({"id": resource.vendor_id})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    
    if vendor.get('status') != 'approved':
        raise HTTPException(status_code=400, detail="Vendor must be approved")
    
    # Check if resource duration is within contract duration
    contract_end = contract.get('end_date')
    if isinstance(contract_end, str):
        contract_end = datetime.fromisoformat(contract_end)
    
    # Ensure both datetimes are timezone-aware for comparison
    resource_end = resource.end_date
    if resource_end.tzinfo is None:
        resource_end = resource_end.replace(tzinfo=timezone.utc)
    if contract_end.tzinfo is None:
        contract_end = contract_end.replace(tzinfo=timezone.utc)
    
    if resource_end > contract_end:
        raise HTTPException(status_code=400, detail="Resource end date cannot exceed contract end date")
    
    # Generate resource number
    year = datetime.now(timezone.utc).strftime('%y')
    count = await db.resources.count_documents({}) + 1
    resource.resource_number = f"RES-{year}-{count:04d}"
    
    # Populate contract and vendor info
    resource.contract_name = contract.get('title')
    resource.scope = contract.get('sow')
    resource.sla = contract.get('sla')
    resource.contract_duration = f"{contract.get('start_date')} to {contract.get('end_date')}"
    resource.vendor_name = vendor.get('name_english') or vendor.get('commercial_name')
    
    resource.created_by = user.id
    resource_dict = resource.model_dump()
    
    await db.resources.insert_one(resource_dict)
    
    return {
        "message": "Resource registered successfully",
        "resource_number": resource.resource_number
    }

@api_router.get("/resources")
async def get_resources(request: Request, status: Optional[str] = None):
    """Get all resources"""
    await require_role(request, [UserRole.PROCUREMENT_OFFICER, UserRole.SYSTEM_ADMIN, UserRole.PD_OFFICER, UserRole.ADMIN])
    
    query = {}
    if status:
        query["status"] = status
    
    resources = await db.resources.find(query, {"_id": 0}).to_list(1000)
    
    # Check for expired resources and update status
    now = datetime.now(timezone.utc)
    for resource in resources:
        # Convert dates for comparison if needed
        end_date = resource.get('end_date')
        if isinstance(end_date, str):
            end_date = datetime.fromisoformat(end_date)
        
        # Ensure timezone-aware for comparison
        if end_date and end_date.tzinfo is None:
            end_date = end_date.replace(tzinfo=timezone.utc)
        
        # Auto-terminate if end_date passed
        if end_date and end_date < now and resource.get('status') == 'active':
            await db.resources.update_one(
                {"id": resource['id']},
                {"$set": {"status": ResourceStatus.INACTIVE.value}}
            )
            resource['status'] = ResourceStatus.INACTIVE.value
        
        # Convert datetime objects back to strings for JSON serialization
        if isinstance(resource.get('start_date'), datetime):
            resource['start_date'] = resource['start_date'].isoformat()
        if isinstance(resource.get('end_date'), datetime):
            resource['end_date'] = resource['end_date'].isoformat()
        if isinstance(resource.get('created_at'), datetime):
            resource['created_at'] = resource['created_at'].isoformat()
        if isinstance(resource.get('updated_at'), datetime):
            resource['updated_at'] = resource['updated_at'].isoformat()
    
    return resources

@api_router.get("/resources/{resource_id}")
async def get_resource(resource_id: str, request: Request):
    """Get resource by ID"""
    await require_role(request, [UserRole.PROCUREMENT_OFFICER, UserRole.SYSTEM_ADMIN, UserRole.PD_OFFICER, UserRole.ADMIN])
    
    resource = await db.resources.find_one({"id": resource_id}, {"_id": 0})
    if not resource:
        raise HTTPException(status_code=404, detail="Resource not found")
    
    return resource

@api_router.put("/resources/{resource_id}")
async def update_resource(resource_id: str, resource_data: dict, request: Request):
    """Update resource details"""
    user = await require_role(request, [UserRole.PROCUREMENT_OFFICER, UserRole.SYSTEM_ADMIN, UserRole.PD_OFFICER, UserRole.ADMIN])
    
    resource = await db.resources.find_one({"id": resource_id})
    if not resource:
        raise HTTPException(status_code=404, detail="Resource not found")
    
    # Only allow updating certain fields
    allowed_fields = ['name', 'nationality', 'id_number', 'education_qualification', 
                     'years_of_experience', 'scope_of_work', 'has_relatives', 'relatives',
                     'access_development', 'access_production', 'access_uat', 'start_date', 'end_date']
    update_data = {k: v for k, v in resource_data.items() if k in allowed_fields}
    
    # Convert date strings to ISO format if provided
    if 'start_date' in update_data and update_data['start_date']:
        if not isinstance(update_data['start_date'], str):
            update_data['start_date'] = update_data['start_date'].isoformat()
    if 'end_date' in update_data and update_data['end_date']:
        if not isinstance(update_data['end_date'], str):
            update_data['end_date'] = update_data['end_date'].isoformat()
    
    # Check if end_date is in the past and auto-update status
    if 'end_date' in update_data:
        end_date = update_data['end_date']
        if isinstance(end_date, str):
            end_date = datetime.fromisoformat(end_date)
        if end_date.tzinfo is None:
            end_date = end_date.replace(tzinfo=timezone.utc)
        
        now = datetime.now(timezone.utc)
        if end_date < now and resource.get('status') == ResourceStatus.ACTIVE.value:
            update_data['status'] = ResourceStatus.INACTIVE.value
    
    update_data['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    await db.resources.update_one(
        {"id": resource_id},
        {"$set": update_data}
    )
    
    return {"message": "Resource updated successfully"}

@api_router.post("/resources/{resource_id}/terminate")
async def terminate_resource(resource_id: str, request: Request, reason: str = "Manual termination"):
    """Terminate a resource"""
    user = await require_role(request, [UserRole.PROCUREMENT_OFFICER, UserRole.SYSTEM_ADMIN, UserRole.PD_OFFICER, UserRole.ADMIN])
    
    resource = await db.resources.find_one({"id": resource_id})
    if not resource:
        raise HTTPException(status_code=404, detail="Resource not found")
    
    await db.resources.update_one(
        {"id": resource_id},
        {"$set": {
            "status": ResourceStatus.TERMINATED.value,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "Resource terminated successfully"}

# ==================== DASHBOARD ENDPOINTS ====================
@api_router.get("/dashboard/stats")
async def get_dashboard_summary_stats(request: Request):
    """Get dashboard summary statistics"""
    await require_role(request, [UserRole.PROCUREMENT_OFFICER, UserRole.PROJECT_MANAGER, UserRole.SYSTEM_ADMIN])
    
    total_vendors = await db.vendors.count_documents({})
    approved_vendors = await db.vendors.count_documents({"status": VendorStatus.APPROVED.value})
    pending_vendors = await db.vendors.count_documents({"status": VendorStatus.PENDING.value})
    
    total_tenders = await db.tenders.count_documents({})
    active_tenders = await db.tenders.count_documents({"status": TenderStatus.PUBLISHED.value})
    
    total_contracts = await db.contracts.count_documents({})
    active_contracts = await db.contracts.count_documents({"status": ContractStatus.ACTIVE.value})
    
    total_invoices = await db.invoices.count_documents({})
    pending_invoices = await db.invoices.count_documents({"status": InvoiceStatus.PENDING.value})
    approved_invoices = await db.invoices.count_documents({"status": InvoiceStatus.APPROVED.value})
    
    return {
        "vendors": {
            "total": total_vendors,
            "approved": approved_vendors,
            "pending": pending_vendors
        },
        "tenders": {
            "total": total_tenders,
            "active": active_tenders
        },
        "contracts": {
            "total": total_contracts,
            "active": active_contracts
        },
        "invoices": {
            "total": total_invoices,
            "pending": pending_invoices,
            "approved": approved_invoices
        }
    }

@api_router.get("/dashboard/alerts")
async def get_dashboard_alerts(request: Request):
    """Get dashboard alerts"""
    await require_role(request, [UserRole.PROCUREMENT_OFFICER, UserRole.PROJECT_MANAGER, UserRole.SYSTEM_ADMIN])
    
    alerts = []
    
    # Pending vendor approvals
    pending_vendors_count = await db.vendors.count_documents({"status": VendorStatus.PENDING.value})
    if pending_vendors_count > 0:
        alerts.append({
            "type": "pending_approval",
            "message": f"{pending_vendors_count} vendor(s) pending approval",
            "count": pending_vendors_count
        })
    
    # Expiring contracts (next 30 days)
    expiry_date = datetime.now(timezone.utc) + timedelta(days=30)
    expiring_contracts = await db.contracts.count_documents({
        "end_date": {
            "$gte": datetime.now(timezone.utc).isoformat(),
            "$lte": expiry_date.isoformat()
        },
        "status": ContractStatus.ACTIVE.value
    })
    if expiring_contracts > 0:
        alerts.append({
            "type": "expiring_contract",
            "message": f"{expiring_contracts} contract(s) expiring in 30 days",
            "count": expiring_contracts
        })
    
    # Pending invoices
    pending_invoices_count = await db.invoices.count_documents({"status": InvoiceStatus.PENDING.value})
    if pending_invoices_count > 0:
        alerts.append({
            "type": "pending_invoice",
            "message": f"{pending_invoices_count} invoice(s) pending verification",
            "count": pending_invoices_count
        })
    
    return alerts

@api_router.get("/notifications")
async def get_notifications(request: Request):
    """Get user notifications"""
    user = await require_auth(request)
    
    notifications = await db.notifications.find({"user_id": user.id}).sort("created_at", -1).limit(50).to_list(50)
    
    for notif in notifications:
        if isinstance(notif.get('created_at'), str):
            notif['created_at'] = datetime.fromisoformat(notif['created_at'])
    
    return notifications

@api_router.put("/notifications/{notification_id}/read")
async def mark_notification_read(notification_id: str, request: Request):
    """Mark notification as read"""
    user = await require_auth(request)
    
    result = await db.notifications.update_one(
        {"id": notification_id, "user_id": user.id},
        {"$set": {"read": True}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Notification not found")
    
    return {"message": "Notification marked as read"}

# ==================== AI ENDPOINTS ====================
@api_router.post("/ai/analyze-vendor")
async def ai_analyze_vendor(vendor_data: dict):
    """AI analyzes vendor data and suggests risk scores"""
    try:
        result = await analyze_vendor_scoring(vendor_data)
        return result
    except Exception as e:
        return {
            "error": str(e),
            "risk_category": "medium",
            "risk_score": 50,
            "reasoning": "AI analysis unavailable"
        }

@api_router.post("/ai/analyze-tender-proposal")
async def ai_analyze_tender(tender_data: dict, proposal_data: dict):
    """AI analyzes tender proposal and suggests scores"""
    try:
        result = await analyze_tender_proposal(tender_data, proposal_data)
        return result
    except Exception as e:
        return {
            "error": str(e),
            "overall_score": 70,
            "recommendation": "Manual Review Required"
        }

@api_router.post("/ai/classify-contract")
async def ai_classify_contract(data: dict):
    """AI classifies contract type (outsourcing, cloud, NOC)"""
    try:
        description = data.get('description', '')
        title = data.get('title', '')
        result = await analyze_contract_classification(description, title)
        return result
    except Exception as e:
        return {
            "error": str(e),
            "outsourcing_classification": "none",
            "is_noc_required": False,
            "reasoning": "AI analysis unavailable"
        }

@api_router.post("/ai/analyze-po-item")
async def ai_analyze_po_item(data: dict):
    """AI analyzes PO item and suggests checkboxes"""
    try:
        description = data.get('description', '')
        result = await analyze_po_items(description)
        return result
    except Exception as e:
        return {
            "error": str(e),
            "requires_contract": False,
            "reasoning": "AI analysis unavailable"
        }

@api_router.post("/ai/match-invoice-milestone")
async def ai_match_invoice_milestone(data: dict):
    """AI matches invoice to contract milestones"""
    try:
        description = data.get('description', '')
        milestones = data.get('milestones', [])
        result = await match_invoice_to_milestone(description, milestones)
        return result
    except Exception as e:
        return {
            "error": str(e),
            "matched_milestone_name": None,
            "reasoning": "AI analysis unavailable"
        }

# ==================== BASIC ENDPOINTS ====================
# ==================== EXPORT ENDPOINTS ====================

@api_router.get("/export/vendors")
async def export_vendors(request: Request):
    """Export all vendors to Excel with complete due diligence data"""
    await require_auth(request)
    
    vendors = await db.vendors.find({}, {"_id": 0}).to_list(1000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Vendors"
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Define comprehensive headers with all actual fields
    headers = [
        "ID", "Name (English)", "Commercial Name", "Entity Type", 
        "VAT Number", "Unified Number", "CR Number", "CR Expiry Date", "CR Country/City",
        "License Number", "License Expiry Date",
        "Activity Description", "Number of Employees",
        "Country", "City", "District", "Street", "Building No", 
        "Representative Name", "Representative Email", "Representative Mobile", 
        "Representative Designation", "Representative Nationality", "Representative ID Type", "Representative ID Number",
        "Email", "Mobile", "Landline", "Fax",
        "Bank Name", "IBAN", "Bank Branch", "Bank Country", "Bank Account Name", "SWIFT Code", "Currency",
        "Status", "Risk Category", "Risk Score",
        "Created At", "Updated At"
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Write data
    for row_idx, vendor in enumerate(vendors, 2):
        ws.cell(row=row_idx, column=1, value=vendor.get("id", ""))
        ws.cell(row=row_idx, column=2, value=vendor.get("name_english", ""))
        ws.cell(row=row_idx, column=3, value=vendor.get("commercial_name", ""))
        ws.cell(row=row_idx, column=4, value=vendor.get("entity_type", ""))
        ws.cell(row=row_idx, column=5, value=vendor.get("vat_number", ""))
        ws.cell(row=row_idx, column=6, value=vendor.get("unified_number", ""))
        ws.cell(row=row_idx, column=7, value=vendor.get("cr_number", ""))
        ws.cell(row=row_idx, column=8, value=str(vendor.get("cr_expiry_date", "")))
        ws.cell(row=row_idx, column=9, value=vendor.get("cr_country_city", ""))
        ws.cell(row=row_idx, column=10, value=vendor.get("license_number", ""))
        ws.cell(row=row_idx, column=11, value=str(vendor.get("license_expiry_date", "")))
        ws.cell(row=row_idx, column=12, value=vendor.get("activity_description", ""))
        ws.cell(row=row_idx, column=13, value=vendor.get("number_of_employees", ""))
        ws.cell(row=row_idx, column=14, value=vendor.get("country", ""))
        ws.cell(row=row_idx, column=15, value=vendor.get("city", ""))
        ws.cell(row=row_idx, column=16, value=vendor.get("district", ""))
        ws.cell(row=row_idx, column=17, value=vendor.get("street", ""))
        ws.cell(row=row_idx, column=18, value=vendor.get("building_no", ""))
        ws.cell(row=row_idx, column=19, value=vendor.get("representative_name", ""))
        ws.cell(row=row_idx, column=20, value=vendor.get("representative_email", ""))
        ws.cell(row=row_idx, column=21, value=vendor.get("representative_mobile", ""))
        ws.cell(row=row_idx, column=22, value=vendor.get("representative_designation", ""))
        ws.cell(row=row_idx, column=23, value=vendor.get("representative_nationality", ""))
        ws.cell(row=row_idx, column=24, value=vendor.get("representative_id_type", ""))
        ws.cell(row=row_idx, column=25, value=vendor.get("representative_id_number", ""))
        ws.cell(row=row_idx, column=26, value=vendor.get("email", ""))
        ws.cell(row=row_idx, column=27, value=vendor.get("mobile", ""))
        ws.cell(row=row_idx, column=28, value=vendor.get("landline", ""))
        ws.cell(row=row_idx, column=29, value=vendor.get("fax", ""))
        ws.cell(row=row_idx, column=30, value=vendor.get("bank_name", ""))
        ws.cell(row=row_idx, column=31, value=vendor.get("iban", ""))
        ws.cell(row=row_idx, column=32, value=vendor.get("bank_branch", ""))
        ws.cell(row=row_idx, column=33, value=vendor.get("bank_country", ""))
        ws.cell(row=row_idx, column=34, value=vendor.get("bank_account_name", ""))
        ws.cell(row=row_idx, column=35, value=vendor.get("swift_code", ""))
        ws.cell(row=row_idx, column=36, value=vendor.get("currency", ""))
        ws.cell(row=row_idx, column=37, value=vendor.get("status", ""))
        ws.cell(row=row_idx, column=38, value=vendor.get("risk_category", ""))
        ws.cell(row=row_idx, column=39, value=vendor.get("risk_score", ""))
        ws.cell(row=row_idx, column=40, value=str(vendor.get("created_at", "")))
        ws.cell(row=row_idx, column=41, value=str(vendor.get("updated_at", "")))
    
    # Auto-size columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=vendors_export.xlsx"}
    )

@api_router.get("/export/contracts")
async def export_contracts(request: Request):
    """Export all contracts with milestones to Excel"""
    await require_auth(request)
    
    contracts = await db.contracts.find({}, {"_id": 0}).to_list(1000)
    
    wb = Workbook()
    
    # Sheet 1: Contracts
    ws_contracts = wb.active
    ws_contracts.title = "Contracts"
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    contract_headers = ["ID", "Contract Number", "Title", "Tender ID", "Vendor ID", 
                        "Status", "Value", "Start Date", "End Date", "Duration (months)",
                        "Statement of Work", "SLA", 
                        "Classification", "NOC Required", "Data Access", "Subcontracting",
                        "Is Outsourcing", "Created By", "Approved By",
                        "Created At", "Updated At"]
    
    for col, header in enumerate(contract_headers, 1):
        cell = ws_contracts.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    for row_idx, contract in enumerate(contracts, 2):
        ws_contracts.cell(row=row_idx, column=1, value=contract.get("id", ""))
        ws_contracts.cell(row=row_idx, column=2, value=contract.get("contract_number", ""))
        ws_contracts.cell(row=row_idx, column=3, value=contract.get("title", ""))
        ws_contracts.cell(row=row_idx, column=4, value=contract.get("tender_id", ""))
        ws_contracts.cell(row=row_idx, column=5, value=contract.get("vendor_id", ""))
        ws_contracts.cell(row=row_idx, column=6, value=contract.get("status", ""))
        ws_contracts.cell(row=row_idx, column=7, value=contract.get("value", 0))
        ws_contracts.cell(row=row_idx, column=8, value=str(contract.get("start_date", "")))
        ws_contracts.cell(row=row_idx, column=9, value=str(contract.get("end_date", "")))
        ws_contracts.cell(row=row_idx, column=10, value=contract.get("duration_months", ""))
        ws_contracts.cell(row=row_idx, column=11, value=contract.get("sow", ""))
        ws_contracts.cell(row=row_idx, column=12, value=contract.get("sla", ""))
        ws_contracts.cell(row=row_idx, column=13, value=contract.get("outsourcing_classification", ""))
        ws_contracts.cell(row=row_idx, column=14, value=str(contract.get("is_noc", False)))
        ws_contracts.cell(row=row_idx, column=15, value=str(contract.get("involves_data_access", False)))
        ws_contracts.cell(row=row_idx, column=16, value=str(contract.get("involves_subcontracting", False)))
        ws_contracts.cell(row=row_idx, column=17, value=str(contract.get("is_outsourcing", False)))
        ws_contracts.cell(row=row_idx, column=18, value=contract.get("created_by", ""))
        ws_contracts.cell(row=row_idx, column=19, value=contract.get("approved_by", ""))
        ws_contracts.cell(row=row_idx, column=20, value=str(contract.get("created_at", "")))
        ws_contracts.cell(row=row_idx, column=21, value=str(contract.get("updated_at", "")))
    
    # Auto-size
    for col in ws_contracts.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws_contracts.column_dimensions[column].width = min(max_length + 2, 50)
    
    # Sheet 2: Milestones
    ws_milestones = wb.create_sheet("Milestones")
    milestone_headers = ["Contract ID", "Contract Number", "Milestone Name", "Description",
                         "Due Date", "Payment Percentage", "Amount", "Status", "Completed Date"]
    
    for col, header in enumerate(milestone_headers, 1):
        cell = ws_milestones.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    milestone_row = 2
    for contract in contracts:
        milestones = contract.get("milestones", [])
        for milestone in milestones:
            ws_milestones.cell(row=milestone_row, column=1, value=contract.get("id", ""))
            ws_milestones.cell(row=milestone_row, column=2, value=contract.get("contract_number", ""))
            ws_milestones.cell(row=milestone_row, column=3, value=milestone.get("name", ""))
            ws_milestones.cell(row=milestone_row, column=4, value=milestone.get("description", ""))
            ws_milestones.cell(row=milestone_row, column=5, value=str(milestone.get("due_date", "")))
            ws_milestones.cell(row=milestone_row, column=6, value=milestone.get("payment_percentage", 0))
            ws_milestones.cell(row=milestone_row, column=7, value=milestone.get("amount", 0))
            ws_milestones.cell(row=milestone_row, column=8, value=milestone.get("status", ""))
            ws_milestones.cell(row=milestone_row, column=9, value=str(milestone.get("completed_date", "")))
            milestone_row += 1
    
    # Auto-size
    for col in ws_milestones.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws_milestones.column_dimensions[column].width = min(max_length + 2, 50)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=contracts_export.xlsx"}
    )

@api_router.get("/export/tenders")
async def export_tenders(request: Request):
    """Export all tenders with proposals and evaluations to Excel"""
    await require_auth(request)
    
    tenders = await db.tenders.find({}, {"_id": 0}).to_list(1000)
    # Fetch all proposals separately
    all_proposals = await db.proposals.find({}, {"_id": 0}).to_list(5000)
    
    wb = Workbook()
    
    # Sheet 1: Tenders
    ws_tenders = wb.active
    ws_tenders.title = "Tenders"
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    tender_headers = ["ID", "Tender Number", "Title", "Description", "Status", "Budget", 
                      "Deadline", "Requirements", "Published Date", "Closing Date", 
                      "Created At", "Updated At"]
    
    for col, header in enumerate(tender_headers, 1):
        cell = ws_tenders.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    for row_idx, tender in enumerate(tenders, 2):
        ws_tenders.cell(row=row_idx, column=1, value=tender.get("id", ""))
        ws_tenders.cell(row=row_idx, column=2, value=tender.get("tender_number", ""))
        ws_tenders.cell(row=row_idx, column=3, value=tender.get("title", ""))
        ws_tenders.cell(row=row_idx, column=4, value=tender.get("description", ""))
        ws_tenders.cell(row=row_idx, column=5, value=tender.get("status", ""))
        ws_tenders.cell(row=row_idx, column=6, value=tender.get("budget", 0))
        ws_tenders.cell(row=row_idx, column=7, value=str(tender.get("deadline", "")))
        ws_tenders.cell(row=row_idx, column=8, value=tender.get("requirements", ""))
        ws_tenders.cell(row=row_idx, column=9, value=str(tender.get("published_date", "")))
        ws_tenders.cell(row=row_idx, column=10, value=str(tender.get("closing_date", "")))
        ws_tenders.cell(row=row_idx, column=11, value=str(tender.get("created_at", "")))
        ws_tenders.cell(row=row_idx, column=12, value=str(tender.get("updated_at", "")))
    
    # Auto-size columns
    for col in ws_tenders.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws_tenders.column_dimensions[column].width = min(max_length + 2, 50)
    
    # Sheet 2: Proposals
    ws_proposals = wb.create_sheet("Proposals")
    proposal_headers = ["Proposal ID", "Tender ID", "Tender Number", "Vendor ID", "Vendor Name",
                        "Proposed Price", "Technical Approach", "Delivery Time", "Status",
                        "Submitted At", "Updated At"]
    
    for col, header in enumerate(proposal_headers, 1):
        cell = ws_proposals.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    proposal_row = 2
    for proposal in all_proposals:
        # Find tender number
        tender_id = proposal.get("tender_id", "")
        tender = next((t for t in tenders if t.get("id") == tender_id), {})
        
        ws_proposals.cell(row=proposal_row, column=1, value=proposal.get("id", ""))
        ws_proposals.cell(row=proposal_row, column=2, value=tender_id)
        ws_proposals.cell(row=proposal_row, column=3, value=tender.get("title", ""))
        ws_proposals.cell(row=proposal_row, column=4, value=proposal.get("vendor_id", ""))
        ws_proposals.cell(row=proposal_row, column=5, value=proposal.get("vendor_name", ""))
        ws_proposals.cell(row=proposal_row, column=6, value=proposal.get("proposed_price", 0))
        ws_proposals.cell(row=proposal_row, column=7, value=proposal.get("technical_approach", ""))
        ws_proposals.cell(row=proposal_row, column=8, value=proposal.get("delivery_time", ""))
        ws_proposals.cell(row=proposal_row, column=9, value=proposal.get("status", ""))
        ws_proposals.cell(row=proposal_row, column=10, value=str(proposal.get("created_at", "")))
        ws_proposals.cell(row=proposal_row, column=11, value=str(proposal.get("updated_at", "")))
        proposal_row += 1
    
    # Auto-size
    for col in ws_proposals.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws_proposals.column_dimensions[column].width = min(max_length + 2, 50)
    
    # Sheet 3: Evaluations
    ws_evaluations = wb.create_sheet("Evaluations")
    eval_headers = ["Evaluation ID", "Tender ID", "Proposal ID", "Vendor Name",
                    "Reliability Score", "Delivery Score", "Technical Score", 
                    "Cost Score", "Meets Requirements", "Total Score",
                    "Evaluated By", "Evaluated At"]
    
    for col, header in enumerate(eval_headers, 1):
        cell = ws_evaluations.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    eval_row = 2
    for proposal in all_proposals:
        evaluation = proposal.get("evaluation", {})
        if evaluation:
            tender_id = proposal.get("tender_id", "")
            tender = next((t for t in tenders if t.get("id") == tender_id), {})
            
            ws_evaluations.cell(row=eval_row, column=1, value=evaluation.get("id", ""))
            ws_evaluations.cell(row=eval_row, column=2, value=tender_id)
            ws_evaluations.cell(row=eval_row, column=3, value=proposal.get("id", ""))
            ws_evaluations.cell(row=eval_row, column=4, value=proposal.get("vendor_name", ""))
            ws_evaluations.cell(row=eval_row, column=5, value=evaluation.get("vendor_reliability_stability", ""))
            ws_evaluations.cell(row=eval_row, column=6, value=evaluation.get("delivery_warranty_backup", ""))
            ws_evaluations.cell(row=eval_row, column=7, value=evaluation.get("technical_experience", ""))
            ws_evaluations.cell(row=eval_row, column=8, value=evaluation.get("cost_score", ""))
            ws_evaluations.cell(row=eval_row, column=9, value=evaluation.get("meets_requirements", ""))
            ws_evaluations.cell(row=eval_row, column=10, value=evaluation.get("total_score", ""))
            ws_evaluations.cell(row=eval_row, column=11, value=evaluation.get("evaluated_by", ""))
            ws_evaluations.cell(row=eval_row, column=12, value=str(evaluation.get("evaluated_at", "")))
            eval_row += 1
    
    # Auto-size
    for col in ws_evaluations.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws_evaluations.column_dimensions[column].width = min(max_length + 2, 50)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=tenders_export.xlsx"}
    )

@api_router.get("/export/invoices")
async def export_invoices(request: Request):
    """Export all invoices with complete details to Excel"""
    await require_auth(request)
    
    invoices = await db.invoices.find({}, {"_id": 0}).to_list(1000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    headers = ["ID", "Invoice Number", "Vendor ID", "Contract ID", "PO ID", "Amount", 
               "Status", "Description", "Issue Date", "Due Date", "Payment Date",
               "Tax Amount", "Discount", "Net Amount",
               "Milestone", "Payment Method", "Notes",
               "Approved By", "Created At", "Updated At"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    for row_idx, invoice in enumerate(invoices, 2):
        ws.cell(row=row_idx, column=1, value=invoice.get("id", ""))
        ws.cell(row=row_idx, column=2, value=invoice.get("invoice_number", ""))
        ws.cell(row=row_idx, column=3, value=invoice.get("vendor_id", ""))
        ws.cell(row=row_idx, column=4, value=invoice.get("contract_id", ""))
        ws.cell(row=row_idx, column=5, value=invoice.get("po_id", ""))
        ws.cell(row=row_idx, column=6, value=invoice.get("amount", 0))
        ws.cell(row=row_idx, column=7, value=invoice.get("status", ""))
        ws.cell(row=row_idx, column=8, value=invoice.get("description", ""))
        ws.cell(row=row_idx, column=9, value=str(invoice.get("issue_date", "")))
        ws.cell(row=row_idx, column=10, value=str(invoice.get("due_date", "")))
        ws.cell(row=row_idx, column=11, value=str(invoice.get("payment_date", "")))
        ws.cell(row=row_idx, column=12, value=invoice.get("tax_amount", 0))
        ws.cell(row=row_idx, column=13, value=invoice.get("discount", 0))
        ws.cell(row=row_idx, column=14, value=invoice.get("net_amount", 0))
        ws.cell(row=row_idx, column=15, value=invoice.get("milestone", ""))
        ws.cell(row=row_idx, column=16, value=invoice.get("payment_method", ""))
        ws.cell(row=row_idx, column=17, value=invoice.get("notes", ""))
        ws.cell(row=row_idx, column=18, value=invoice.get("approved_by", ""))
        ws.cell(row=row_idx, column=19, value=str(invoice.get("created_at", "")))
        ws.cell(row=row_idx, column=20, value=str(invoice.get("updated_at", "")))
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=invoices_export.xlsx"}
    )

@api_router.get("/export/purchase-orders")
async def export_purchase_orders(request: Request):
    """Export all purchase orders with line items to Excel"""
    await require_auth(request)
    
    pos = await db.purchase_orders.find({}, {"_id": 0}).to_list(1000)
    
    wb = Workbook()
    
    # Sheet 1: Purchase Orders
    ws_pos = wb.active
    ws_pos.title = "Purchase Orders"
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    po_headers = ["ID", "PO Number", "Vendor ID", "Tender ID", "Status", 
                  "Total Value", "Delivery Location", "Delivery Date",
                  "Payment Terms", "Notes", "Created By", "Approved By",
                  "Created At", "Updated At"]
    
    for col, header in enumerate(po_headers, 1):
        cell = ws_pos.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    for row_idx, po in enumerate(pos, 2):
        ws_pos.cell(row=row_idx, column=1, value=po.get("id", ""))
        ws_pos.cell(row=row_idx, column=2, value=po.get("po_number", ""))
        ws_pos.cell(row=row_idx, column=3, value=po.get("vendor_id", ""))
        ws_pos.cell(row=row_idx, column=4, value=po.get("tender_id", ""))
        ws_pos.cell(row=row_idx, column=5, value=po.get("status", ""))
        ws_pos.cell(row=row_idx, column=6, value=po.get("total_value", 0))
        ws_pos.cell(row=row_idx, column=7, value=po.get("delivery_location", ""))
        ws_pos.cell(row=row_idx, column=8, value=str(po.get("delivery_date", "")))
        ws_pos.cell(row=row_idx, column=9, value=po.get("payment_terms", ""))
        ws_pos.cell(row=row_idx, column=10, value=po.get("notes", ""))
        ws_pos.cell(row=row_idx, column=11, value=po.get("created_by", ""))
        ws_pos.cell(row=row_idx, column=12, value=po.get("approved_by", ""))
        ws_pos.cell(row=row_idx, column=13, value=str(po.get("created_at", "")))
        ws_pos.cell(row=row_idx, column=14, value=str(po.get("updated_at", "")))
    
    # Auto-size
    for col in ws_pos.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws_pos.column_dimensions[column].width = min(max_length + 2, 50)
    
    # Sheet 2: PO Items
    ws_items = wb.create_sheet("PO Items")
    item_headers = ["PO ID", "PO Number", "Item Name", "Description", 
                    "Quantity", "Unit Price", "Total", "Unit", "Category"]
    
    for col, header in enumerate(item_headers, 1):
        cell = ws_items.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    item_row = 2
    for po in pos:
        items = po.get("items", [])
        for item in items:
            ws_items.cell(row=item_row, column=1, value=po.get("id", ""))
            ws_items.cell(row=item_row, column=2, value=po.get("po_number", ""))
            ws_items.cell(row=item_row, column=3, value=item.get("name", ""))
            ws_items.cell(row=item_row, column=4, value=item.get("description", ""))
            ws_items.cell(row=item_row, column=5, value=item.get("quantity", 0))
            ws_items.cell(row=item_row, column=6, value=item.get("price", 0))
            ws_items.cell(row=item_row, column=7, value=item.get("total", 0))
            ws_items.cell(row=item_row, column=8, value=item.get("unit", ""))
            ws_items.cell(row=item_row, column=9, value=item.get("category", ""))
            item_row += 1
    
    # Auto-size
    for col in ws_items.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws_items.column_dimensions[column].width = min(max_length + 2, 50)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=purchase_orders_export.xlsx"}
    )

@api_router.get("/export/resources")
async def export_resources(request: Request):
    """Export all resources with complete details to Excel"""
    await require_auth(request)
    
    resources = await db.resources.find({}, {"_id": 0}).to_list(1000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Resources"
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    headers = ["ID", "Name", "Resource Type", "Vendor ID", "Contract ID", 
               "Location", "Location Type", "Status", "Position/Role", 
               "Department", "Start Date", "End Date", "Cost", 
               "Qualifications", "Experience", "Certifications",
               "Contact Email", "Contact Phone", "Notes",
               "Created At", "Updated At"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    for row_idx, resource in enumerate(resources, 2):
        ws.cell(row=row_idx, column=1, value=resource.get("id", ""))
        ws.cell(row=row_idx, column=2, value=resource.get("name", ""))
        ws.cell(row=row_idx, column=3, value=resource.get("resource_type", ""))
        ws.cell(row=row_idx, column=4, value=resource.get("vendor_id", ""))
        ws.cell(row=row_idx, column=5, value=resource.get("contract_id", ""))
        ws.cell(row=row_idx, column=6, value=resource.get("location", ""))
        ws.cell(row=row_idx, column=7, value=resource.get("location_type", ""))
        ws.cell(row=row_idx, column=8, value=resource.get("status", ""))
        ws.cell(row=row_idx, column=9, value=resource.get("position", ""))
        ws.cell(row=row_idx, column=10, value=resource.get("department", ""))
        ws.cell(row=row_idx, column=11, value=str(resource.get("start_date", "")))
        ws.cell(row=row_idx, column=12, value=str(resource.get("end_date", "")))
        ws.cell(row=row_idx, column=13, value=resource.get("cost", 0))
        ws.cell(row=row_idx, column=14, value=resource.get("qualifications", ""))
        ws.cell(row=row_idx, column=15, value=resource.get("experience", ""))
        ws.cell(row=row_idx, column=16, value=resource.get("certifications", ""))
        ws.cell(row=row_idx, column=17, value=resource.get("email", ""))
        ws.cell(row=row_idx, column=18, value=resource.get("phone", ""))
        ws.cell(row=row_idx, column=19, value=resource.get("notes", ""))
        ws.cell(row=row_idx, column=20, value=str(resource.get("created_at", "")))
        ws.cell(row=row_idx, column=21, value=str(resource.get("updated_at", "")))
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=resources_export.xlsx"}
    )

@api_router.get("/")
async def root():
    return {"message": "Sourcevia Procurement Management System API"}

# App setup moved to end of file after all endpoints are defined


# ==================== FILE UPLOAD ENDPOINTS ====================
from fastapi import File, UploadFile
from pathlib import Path
import shutil

UPLOAD_DIR = Path("/app/backend/uploads")

@api_router.post("/upload/vendor/{vendor_id}")
async def upload_vendor_files(
    vendor_id: str,
    request: Request,
    files: List[UploadFile] = File(...),
    file_type: str = "supporting_documents"  # supporting_documents, due_diligence
):
    """Upload files for vendor (supporting documents, due diligence)"""
    await require_auth(request)
    
    uploaded_files = []
    vendor_dir = UPLOAD_DIR / "vendors" / vendor_id
    vendor_dir.mkdir(parents=True, exist_ok=True)
    
    for file in files:
        # Generate unique filename
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        filename = f"{timestamp}_{file.filename}"
        file_path = vendor_dir / filename
        
        # Save file
        with file_path.open("wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        
        uploaded_files.append({
            "filename": file.filename,
            "stored_filename": filename,
            "file_type": file_type,
            "size": file_path.stat().st_size,
            "uploaded_at": datetime.now(timezone.utc).isoformat()
        })
    
    # Update vendor record with file metadata
    await db.vendors.update_one(
        {"id": vendor_id},
        {"$push": {"attachments": {"$each": uploaded_files}}}
    )
    
    return {"message": f"Uploaded {len(uploaded_files)} files", "files": uploaded_files}

@api_router.post("/upload/tender/{tender_id}")
async def upload_tender_files(
    tender_id: str,
    request: Request,
    files: List[UploadFile] = File(...)
):
    """Upload supporting documents for tender"""
    await require_auth(request)
    
    uploaded_files = []
    tender_dir = UPLOAD_DIR / "tenders" / tender_id
    tender_dir.mkdir(parents=True, exist_ok=True)
    
    for file in files:
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        filename = f"{timestamp}_{file.filename}"
        file_path = tender_dir / filename
        
        with file_path.open("wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        
        uploaded_files.append({
            "filename": file.filename,
            "stored_filename": filename,
            "size": file_path.stat().st_size,
            "uploaded_at": datetime.now(timezone.utc).isoformat()
        })
    
    await db.tenders.update_one(
        {"id": tender_id},
        {"$push": {"attachments": {"$each": uploaded_files}}}
    )
    
    return {"message": f"Uploaded {len(uploaded_files)} files", "files": uploaded_files}

@api_router.post("/upload/proposal/{proposal_id}")
async def upload_proposal_files(
    proposal_id: str,
    request: Request,
    files: List[UploadFile] = File(...)
):
    """Upload supporting documents for proposal"""
    await require_auth(request)
    
    uploaded_files = []
    proposal_dir = UPLOAD_DIR / "proposals" / proposal_id
    proposal_dir.mkdir(parents=True, exist_ok=True)
    
    for file in files:
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        filename = f"{timestamp}_{file.filename}"
        file_path = proposal_dir / filename
        
        with file_path.open("wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        
        uploaded_files.append({
            "filename": file.filename,
            "stored_filename": filename,
            "size": file_path.stat().st_size,
            "uploaded_at": datetime.now(timezone.utc).isoformat()
        })
    
    await db.proposals.update_one(
        {"id": proposal_id},
        {"$push": {"attachments": {"$each": uploaded_files}}}
    )
    
    return {"message": f"Uploaded {len(uploaded_files)} files", "files": uploaded_files}

@api_router.post("/upload/purchase-order/{po_id}")
async def upload_po_files(
    po_id: str,
    request: Request,
    files: List[UploadFile] = File(...),
    file_type: str = "quotation"  # quotation, supporting_documents
):
    """Upload quotations and supporting documents for PO"""
    await require_auth(request)
    
    uploaded_files = []
    po_dir = UPLOAD_DIR / "purchase_orders" / po_id
    po_dir.mkdir(parents=True, exist_ok=True)
    
    for file in files:
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        filename = f"{timestamp}_{file.filename}"
        file_path = po_dir / filename
        
        with file_path.open("wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        
        uploaded_files.append({
            "filename": file.filename,
            "stored_filename": filename,
            "file_type": file_type,
            "size": file_path.stat().st_size,
            "uploaded_at": datetime.now(timezone.utc).isoformat()
        })
    
    await db.purchase_orders.update_one(
        {"id": po_id},
        {"$push": {"attachments": {"$each": uploaded_files}}}
    )
    
    return {"message": f"Uploaded {len(uploaded_files)} files", "files": uploaded_files}

@api_router.post("/upload/invoice/{invoice_id}")
async def upload_invoice_files(
    invoice_id: str,
    request: Request,
    files: List[UploadFile] = File(...)
):
    """Upload invoice files"""
    await require_auth(request)
    
    uploaded_files = []
    invoice_dir = UPLOAD_DIR / "invoices" / invoice_id
    invoice_dir.mkdir(parents=True, exist_ok=True)
    
    for file in files:
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        filename = f"{timestamp}_{file.filename}"
        file_path = invoice_dir / filename
        
        with file_path.open("wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        
        uploaded_files.append({
            "filename": file.filename,
            "stored_filename": filename,
            "size": file_path.stat().st_size,
            "uploaded_at": datetime.now(timezone.utc).isoformat()
        })
    
    await db.invoices.update_one(
        {"id": invoice_id},
        {"$push": {"attachments": {"$each": uploaded_files}}}
    )
    
    return {"message": f"Uploaded {len(uploaded_files)} files", "files": uploaded_files}

@api_router.post("/upload/resource/{resource_id}")
async def upload_resource_files(
    resource_id: str,
    request: Request,
    files: List[UploadFile] = File(...)
):
    """Upload supporting documents for resource"""
    await require_auth(request)
    
    uploaded_files = []
    resource_dir = UPLOAD_DIR / "resources" / resource_id
    resource_dir.mkdir(parents=True, exist_ok=True)
    
    for file in files:
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        filename = f"{timestamp}_{file.filename}"
        file_path = resource_dir / filename
        
        with file_path.open("wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        
        uploaded_files.append({
            "filename": file.filename,
            "stored_filename": filename,
            "size": file_path.stat().st_size,
            "uploaded_at": datetime.now(timezone.utc).isoformat()
        })
    
    await db.resources.update_one(
        {"id": resource_id},
        {"$push": {"attachments": {"$each": uploaded_files}}}
    )
    
    return {"message": f"Uploaded {len(uploaded_files)} files", "files": uploaded_files}


@api_router.post("/upload/contract/{contract_id}")
async def upload_contract_files(
    contract_id: str,
    request: Request,
    files: List[UploadFile] = File(...)
):
    """Upload contract documents"""
    await require_auth(request)
    
    uploaded_files = []
    contract_dir = UPLOAD_DIR / "contracts" / contract_id
    contract_dir.mkdir(parents=True, exist_ok=True)
    
    for file in files:
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        filename = f"{timestamp}_{file.filename}"
        file_path = contract_dir / filename
        
        with file_path.open("wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        
        uploaded_files.append({
            "filename": file.filename,
            "stored_filename": filename,
            "size": file_path.stat().st_size,
            "uploaded_at": datetime.now(timezone.utc).isoformat()
        })
    
    await db.contracts.update_one(
        {"id": contract_id},
        {"$push": {"attachments": {"$each": uploaded_files}}}
    )
    
    return {"message": f"Uploaded {len(uploaded_files)} files", "files": uploaded_files}

@api_router.get("/download/{module}/{entity_id}/{filename}")
async def download_file(module: str, entity_id: str, filename: str, request: Request):
    """Download uploaded file"""
    await require_auth(request)
    
    file_path = UPLOAD_DIR / module / entity_id / filename
    
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    
    from fastapi.responses import FileResponse
    return FileResponse(
        path=file_path,
        filename=filename.split("_", 1)[1] if "_" in filename else filename  # Remove timestamp prefix
    )

# ==================== FACILITIES MANAGEMENT ENDPOINTS ====================

# Buildings
@api_router.get("/buildings")
async def get_buildings(request: Request):
    """Get all buildings"""
    await require_auth(request)
    buildings = await db.buildings.find({}, {"_id": 0}).to_list(1000)
    return buildings

@api_router.post("/buildings")
async def create_building(request: Request, building: Building):
    """Create a new building"""
    await require_auth(request)
    building_dict = building.model_dump()
    await db.buildings.insert_one(building_dict)
    return {"message": "Building created successfully", "building": building_dict}

@api_router.put("/buildings/{building_id}")
async def update_building(building_id: str, request: Request, building: Building):
    """Update building"""
    await require_auth(request)
    building_dict = building.model_dump()
    building_dict["updated_at"] = datetime.now(timezone.utc)
    await db.buildings.update_one({"id": building_id}, {"$set": building_dict})
    return {"message": "Building updated successfully"}

@api_router.delete("/buildings/{building_id}")
async def delete_building(building_id: str, request: Request):
    """Delete building"""
    await require_auth(request)
    await db.buildings.delete_one({"id": building_id})
    return {"message": "Building deleted successfully"}

# Floors
@api_router.get("/floors")
async def get_floors(request: Request, building_id: Optional[str] = None):
    """Get all floors, optionally filtered by building"""
    await require_auth(request)
    query = {}
    if building_id:
        query["building_id"] = building_id
    floors = await db.floors.find(query, {"_id": 0}).to_list(1000)
    return floors

@api_router.post("/floors")
async def create_floor(request: Request, floor: Floor):
    """Create a new floor"""
    await require_auth(request)
    floor_dict = floor.model_dump()
    await db.floors.insert_one(floor_dict)
    return {"message": "Floor created successfully", "floor": floor_dict}

@api_router.put("/floors/{floor_id}")
async def update_floor(floor_id: str, request: Request, floor: Floor):
    """Update floor"""
    await require_auth(request)
    floor_dict = floor.model_dump()
    floor_dict["updated_at"] = datetime.now(timezone.utc)
    await db.floors.update_one({"id": floor_id}, {"$set": floor_dict})
    return {"message": "Floor updated successfully"}

@api_router.delete("/floors/{floor_id}")
async def delete_floor(floor_id: str, request: Request):
    """Delete floor"""
    await require_auth(request)
    await db.floors.delete_one({"id": floor_id})
    return {"message": "Floor deleted successfully"}

# Asset Categories
@api_router.get("/asset-categories")
async def get_asset_categories(request: Request):
    """Get all asset categories"""
    await require_auth(request)
    categories = await db.asset_categories.find({}, {"_id": 0}).to_list(1000)
    return categories

@api_router.post("/asset-categories")
async def create_asset_category(request: Request, category: AssetCategory):
    """Create a new asset category"""
    await require_auth(request)
    category_dict = category.model_dump()
    await db.asset_categories.insert_one(category_dict)
    return {"message": "Asset category created successfully", "category": category_dict}

@api_router.put("/asset-categories/{category_id}")
async def update_asset_category(category_id: str, request: Request, category: AssetCategory):
    """Update asset category"""
    await require_auth(request)
    category_dict = category.model_dump()
    category_dict["updated_at"] = datetime.now(timezone.utc)
    await db.asset_categories.update_one({"id": category_id}, {"$set": category_dict})
    return {"message": "Asset category updated successfully"}

@api_router.delete("/asset-categories/{category_id}")
async def delete_asset_category(category_id: str, request: Request):
    """Delete asset category"""
    await require_auth(request)
    await db.asset_categories.delete_one({"id": category_id})
    return {"message": "Asset category deleted successfully"}

# OSR Categories
@api_router.get("/osr-categories")
async def get_osr_categories(request: Request):
    """Get all OSR categories"""
    await require_auth(request)
    categories = await db.osr_categories.find({}, {"_id": 0}).to_list(1000)
    return categories

@api_router.post("/osr-categories")
async def create_osr_category(request: Request, category_data: dict):
    """Create OSR category"""
    await require_auth(request)
    category_data["id"] = str(uuid.uuid4())
    category_data["created_at"] = datetime.now(timezone.utc)
    await db.osr_categories.insert_one(category_data)
    return {"message": "OSR category created successfully", "category": category_data}

@api_router.put("/osr-categories/{category_id}")
async def update_osr_category(category_id: str, request: Request, category_data: dict):
    """Update OSR category"""
    await require_auth(request)
    category_data["updated_at"] = datetime.now(timezone.utc)
    await db.osr_categories.update_one({"id": category_id}, {"$set": category_data})
    return {"message": "OSR category updated successfully"}

@api_router.delete("/osr-categories/{category_id}")
async def delete_osr_category(category_id: str, request: Request):
    """Delete OSR category"""
    await require_auth(request)
    await db.osr_categories.delete_one({"id": category_id})
    return {"message": "OSR category deleted successfully"}

# Assets
@api_router.get("/assets")
async def get_assets(request: Request):
    """Get all assets"""
    await require_auth(request)
    assets = await db.assets.find({}, {"_id": 0}).to_list(10000)
    
    # Enrich assets with denormalized data
    for asset in assets:
        # Get category name
        if asset.get("category_id"):
            category = await db.asset_categories.find_one({"id": asset["category_id"]}, {"_id": 0})
            if category:
                asset["category_name"] = category.get("name")
        
        # Get building name
        if asset.get("building_id"):
            building = await db.buildings.find_one({"id": asset["building_id"]}, {"_id": 0})
            if building:
                asset["building_name"] = building.get("name")
        
        # Get floor name
        if asset.get("floor_id"):
            floor = await db.floors.find_one({"id": asset["floor_id"]}, {"_id": 0})
            if floor:
                asset["floor_name"] = floor.get("name")
        
        # Get vendor name
        if asset.get("vendor_id"):
            vendor = await db.vendors.find_one({"id": asset["vendor_id"]}, {"_id": 0})
            if vendor:
                asset["vendor_name"] = vendor.get("name_english")
        
        # Get contract number
        if asset.get("contract_id"):
            contract = await db.contracts.find_one({"id": asset["contract_id"]}, {"_id": 0})
            if contract:
                asset["contract_number"] = contract.get("contract_number")
    
    return assets

@api_router.get("/assets/{asset_id}")
async def get_asset(asset_id: str, request: Request):
    """Get single asset"""
    await require_auth(request)
    asset = await db.assets.find_one({"id": asset_id}, {"_id": 0})
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")
    
    # Enrich asset with denormalized data
    if asset.get("category_id"):
        category = await db.asset_categories.find_one({"id": asset["category_id"]}, {"_id": 0})
        if category:
            asset["category_name"] = category.get("name")
    
    if asset.get("building_id"):
        building = await db.buildings.find_one({"id": asset["building_id"]}, {"_id": 0})
        if building:
            asset["building_name"] = building.get("name")
    
    if asset.get("floor_id"):
        floor = await db.floors.find_one({"id": asset["floor_id"]}, {"_id": 0})
        if floor:
            asset["floor_name"] = floor.get("name")
    
    if asset.get("vendor_id"):
        vendor = await db.vendors.find_one({"id": asset["vendor_id"]}, {"_id": 0})
        if vendor:
            asset["vendor_name"] = vendor.get("name_english")
    
    if asset.get("contract_id"):
        contract = await db.contracts.find_one({"id": asset["contract_id"]}, {"_id": 0})
        if contract:
            asset["contract_number"] = contract.get("contract_number")
    
    return asset

@api_router.post("/assets")
async def create_asset(request: Request, asset: Asset):
    """Create a new asset"""
    await require_auth(request)
    
    # Generate asset number
    count = await db.assets.count_documents({})
    asset.asset_number = f"ASSET-{datetime.now().year}-{str(count + 1).zfill(4)}"
    
    # Calculate warranty status
    if asset.warranty_end_date:
        if datetime.now(timezone.utc) <= asset.warranty_end_date:
            asset.warranty_status = "in_warranty"
        else:
            asset.warranty_status = "out_of_warranty"
    
    asset_dict = asset.model_dump()
    asset_dict["created_at"] = datetime.now(timezone.utc)
    
    await db.assets.insert_one(asset_dict)
    
    # Remove _id for response to avoid serialization issues
    response_asset = {k: v for k, v in asset_dict.items() if k != "_id"}
    return {"message": "Asset created successfully", "asset": response_asset}

@api_router.put("/assets/{asset_id}")
async def update_asset(asset_id: str, request: Request, asset: Asset):
    """Update asset"""
    await require_auth(request)
    
    # Recalculate warranty status
    if asset.warranty_end_date:
        if datetime.now(timezone.utc) <= asset.warranty_end_date:
            asset.warranty_status = "in_warranty"
        else:
            asset.warranty_status = "out_of_warranty"
    
    asset_dict = asset.model_dump()
    asset_dict["updated_at"] = datetime.now(timezone.utc)
    
    await db.assets.update_one({"id": asset_id}, {"$set": asset_dict})
    return {"message": "Asset updated successfully"}

@api_router.delete("/assets/{asset_id}")
async def delete_asset(asset_id: str, request: Request):
    """Delete asset"""
    await require_auth(request)
    await db.assets.delete_one({"id": asset_id})
    return {"message": "Asset deleted successfully"}

# OSR (Operating Service Requests)
@api_router.get("/osrs")
async def get_osrs(request: Request):
    """Get all OSRs"""
    await require_auth(request)
    osrs = await db.osr.find({"_id": 0}).to_list(10000)
    return osrs

@api_router.get("/osrs/{osr_id}")
async def get_osr(osr_id: str, request: Request):
    """Get single OSR"""
    await require_auth(request)
    osr = await db.osr.find_one({"id": osr_id}, {"_id": 0})
    if not osr:
        raise HTTPException(status_code=404, detail="OSR not found")
    return osr

@api_router.post("/osrs")
async def create_osr(request: Request, osr: OSR):
    """Create a new OSR"""
    user = await require_auth(request)
    
    # Generate OSR number
    count = await db.osr.count_documents({})
    osr.osr_number = f"OSR-{datetime.now().year}-{str(count + 1).zfill(4)}"
    
    osr.created_by = user.id
    osr.created_by_name = user.name
    
    # If asset-related, fetch asset details
    if osr.request_type == OSRType.ASSET_RELATED and osr.asset_id:
        asset = await db.assets.find_one({"id": osr.asset_id}, {"_id": 0})
        if asset:
            osr.asset_name = asset.get("name")
            osr.asset_warranty_status = asset.get("warranty_status")
            osr.asset_contract_id = asset.get("contract_id")
            osr.asset_contract_number = asset.get("contract_number")
            
            # Auto-suggest AMC vendor if maintenance request
            if osr.category == OSRCategory.MAINTENANCE and asset.get("vendor_id"):
                osr.assigned_to_type = "vendor"
                osr.assigned_to_vendor_id = asset.get("vendor_id")
                osr.assigned_to_vendor_name = asset.get("vendor_name")
    
    osr_dict = osr.model_dump()
    osr_dict["created_at"] = datetime.now(timezone.utc)
    
    await db.osr.insert_one(osr_dict)
    return {"message": "OSR created successfully", "osr": osr_dict}

@api_router.put("/osrs/{osr_id}")
async def update_osr(osr_id: str, request: Request, update_data: dict):
    """Update OSR"""
    await require_auth(request)
    
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    # Get existing OSR
    existing_osr = await db.osr.find_one({"id": osr_id}, {"_id": 0})
    if not existing_osr:
        raise HTTPException(status_code=404, detail="OSR not found")
    
    # If status changed to completed and it's an asset-related OSR
    if (update_data.get("status") == "completed" and 
        existing_osr.get("status") != "completed" and
        existing_osr.get("request_type") == "asset_related" and
        existing_osr.get("asset_id")):
        
        # Update asset's last_maintenance_date
        update_data["closed_date"] = datetime.now(timezone.utc)
        await db.assets.update_one(
            {"id": existing_osr["asset_id"]},
            {"$set": {"last_maintenance_date": update_data["closed_date"]}}
        )
    
    await db.osr.update_one({"id": osr_id}, {"$set": update_data})
    return {"message": "OSR updated successfully"}

@api_router.delete("/osrs/{osr_id}")
async def delete_osr(osr_id: str, request: Request):
    """Delete OSR"""
    await require_auth(request)
    await db.osr.delete_one({"id": osr_id})
    return {"message": "OSR deleted successfully"}

# Seed Master Data
@api_router.get("/facilities/master-data")
async def get_master_data(request: Request):
    """Get all master data for facilities management"""
    await require_auth(request)
    
    buildings = await db.buildings.find({}, {"_id": 0}).to_list(1000)
    floors = await db.floors.find({}, {"_id": 0}).to_list(1000)
    asset_categories = await db.asset_categories.find({}, {"_id": 0}).to_list(1000)
    osr_categories = await db.osr_categories.find({}, {"_id": 0}).to_list(1000)
    
    return {
        "buildings": buildings,
        "floors": floors,
        "asset_categories": asset_categories,
        "osr_categories": osr_categories
    }

@api_router.post("/facilities/seed-data")
async def seed_facilities_data(request: Request):
    """Seed initial master data for facilities management"""
    await require_auth(request)
    
    # Seed Asset Categories
    categories = [
        {"id": str(uuid.uuid4()), "name": "HVAC", "description": "Heating, Ventilation, and Air Conditioning", "is_active": True, "created_at": datetime.now(timezone.utc)},
        {"id": str(uuid.uuid4()), "name": "Electrical", "description": "Electrical Systems and Equipment", "is_active": True, "created_at": datetime.now(timezone.utc)},
        {"id": str(uuid.uuid4()), "name": "Plumbing", "description": "Plumbing Systems", "is_active": True, "created_at": datetime.now(timezone.utc)},
        {"id": str(uuid.uuid4()), "name": "Fire Safety", "description": "Fire Safety Equipment", "is_active": True, "created_at": datetime.now(timezone.utc)},
        {"id": str(uuid.uuid4()), "name": "Furniture", "description": "Office and Facility Furniture", "is_active": True, "created_at": datetime.now(timezone.utc)},
        {"id": str(uuid.uuid4()), "name": "Elevators", "description": "Elevators and Lifts", "is_active": True, "created_at": datetime.now(timezone.utc)},
        {"id": str(uuid.uuid4()), "name": "Vehicles", "description": "Company Vehicles", "is_active": True, "created_at": datetime.now(timezone.utc)},
        {"id": str(uuid.uuid4()), "name": "Other", "description": "Other Assets", "is_active": True, "created_at": datetime.now(timezone.utc)},
    ]
    
    existing_categories = await db.asset_categories.count_documents({})
    if existing_categories == 0:
        await db.asset_categories.insert_many(categories)
    
    # Seed OSR Categories
    osr_cats = [
        {"id": str(uuid.uuid4()), "name": "Maintenance", "description": "Asset maintenance and repairs", "is_active": True, "created_at": datetime.now(timezone.utc)},
        {"id": str(uuid.uuid4()), "name": "Cleaning", "description": "Cleaning services", "is_active": True, "created_at": datetime.now(timezone.utc)},
        {"id": str(uuid.uuid4()), "name": "Relocation", "description": "Moving and relocation services", "is_active": True, "created_at": datetime.now(timezone.utc)},
        {"id": str(uuid.uuid4()), "name": "Installation", "description": "New equipment installation", "is_active": True, "created_at": datetime.now(timezone.utc)},
        {"id": str(uuid.uuid4()), "name": "Inspection", "description": "Safety and quality inspections", "is_active": True, "created_at": datetime.now(timezone.utc)},
        {"id": str(uuid.uuid4()), "name": "Other", "description": "Other service requests", "is_active": True, "created_at": datetime.now(timezone.utc)},
    ]
    
    existing_osr_cats = await db.osr_categories.count_documents({})
    if existing_osr_cats == 0:
        await db.osr_categories.insert_many(osr_cats)
    
    # Seed Sample Buildings
    buildings = [
        {"id": str(uuid.uuid4()), "name": "Main Office", "code": "MO", "address": "123 Main St", "is_active": True, "created_at": datetime.now(timezone.utc)},
        {"id": str(uuid.uuid4()), "name": "Warehouse A", "code": "WH-A", "address": "456 Industrial Rd", "is_active": True, "created_at": datetime.now(timezone.utc)},
    ]
    
    existing_buildings = await db.buildings.count_documents({})
    if existing_buildings == 0:
        await db.buildings.insert_many(buildings)
        
        # Seed Sample Floors for Main Office
        building_id = buildings[0]["id"]
        floors = [
            {"id": str(uuid.uuid4()), "building_id": building_id, "name": "Ground Floor", "number": 0, "is_active": True, "created_at": datetime.now(timezone.utc)},
            {"id": str(uuid.uuid4()), "building_id": building_id, "name": "1st Floor", "number": 1, "is_active": True, "created_at": datetime.now(timezone.utc)},
            {"id": str(uuid.uuid4()), "building_id": building_id, "name": "2nd Floor", "number": 2, "is_active": True, "created_at": datetime.now(timezone.utc)},
        ]
        await db.floors.insert_many(floors)
    
    return {"message": "Seed data created successfully"}

# ==================== APP SETUP ====================
# Configure CORS middleware (must be before including router)
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Include the router in the main app (must be after all endpoints are defined)
app.include_router(api_router)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()

